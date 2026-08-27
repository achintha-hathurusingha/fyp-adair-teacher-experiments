"""TEST18: build the results Excel workbook LOCALLY from the real
eval/training CSVs pulled back from devon (CSV-first / render-locally
policy, established since TEST07-B)."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST18 = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST18 / "results"
STATS_DIR = RESULTS_DIR / "statistics"
DIAG_DIR = RESULTS_DIR / "frequency_diagrams"
OUT_XLSX = RESULTS_DIR / "AdaIR_Component_Ablation.xlsx"
HEADER_FONT = Font(bold=True)
VARIANTS = ["A_baseline", "B_fixed_mask", "C_learned_mask", "D_plus_lh", "E_full"]


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def main():
    overall = pd.read_csv(STATS_DIR / "eval_summary_overall.csv").set_index("variant").loc[VARIANTS].reset_index()
    by_deg = pd.read_csv(STATS_DIR / "eval_summary_by_degradation.csv")
    per_image = pd.read_csv(STATS_DIR / "eval_per_image.csv")

    epoch_dfs = []
    for v in VARIANTS:
        p = RESULTS_DIR / f"epoch_metrics_{v}.csv"
        if p.exists():
            epoch_dfs.append(pd.read_csv(p))
    epoch_df = pd.concat(epoch_dfs, ignore_index=True) if epoch_dfs else pd.DataFrame()
    final_loss = epoch_df.groupby("variant")["train_l1_loss"].last().reindex(VARIANTS).reset_index()
    final_loss.columns = ["variant", "final_epoch_mean_l1"]

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 130
    readme = [
        "TEST18: AdaIR Component Ablation (Paper-Style Retraining) + Frequency-Domain Diagnostics",
        "",
        "Replicates the STRUCTURE of AdaIR's own Table 7 ablation (baseline -> +FMiM fixed mask -> "
        "+FMiM learned MGB -> +L-H -> +L-H+H-L=full), but on the 3-in-1 degradation setting "
        "(dehaze+derain+denoise) at 8 epochs/variant, real held-out test data throughout.",
        "",
        "HEADLINE FINDING (surprising, reported honestly): unlike the paper's own monotonic a->e "
        "PSNR improvement, this reproduction shows B_fixed_mask (28.74dB) OUTPERFORMING E_full "
        "(28.67dB, = the released architecture retrained from scratch), and C_learned_mask "
        "(27.95dB) as the WORST of all 5 variants -- below even the no-AFLB baseline (28.57dB).",
        "",
        "ROOT CAUSE, independently confirmed via the frequency-domain diagrams (see Frequency_Diagrams "
        "sheet and results/frequency_diagrams/*.png): the learned mask (MGB) is DEGENERATE (uniform "
        "zero half-width) at EVERY AFLB position for EVERY variant, at the 256px diagnostic resolution "
        "-- an exact, independent reproduction of TEST01/TEST06-R's original finding on the FROZEN "
        "released checkpoint, this time observed directly during/after fresh training. The learned "
        "mask never becomes spatially selective at practical training/inference resolutions, so the "
        "'adaptive' mechanism the paper credits with a real PSNR gain is not actually functioning as "
        "adaptive in this regime -- plausibly explaining why C (learned-but-degenerate mask) "
        "underperforms B (fixed, at least non-trivial 10x10 mask).",
        "",
        "This RECONCILES with, rather than contradicts, TEST01-TEST06R's frozen-checkpoint null "
        "result: the frequency modules DO execute (confirmed, matching the FMiM/FMoM's presence in "
        "the forward pass), but the specific adaptive-mask mechanism does not behave adaptively at "
        "the resolutions actually used for training or typical inference -- true both for the frozen "
        "released checkpoint (TEST01-06R) and for a freshly-trained model (TEST18).",
        "",
        "SCOPE CAVEATS (see TEST18_PLAN.md and test18_report.md for full detail): single seed per "
        "variant (not 3-seed); 8 epochs on a 10k-image dehaze subsample (not the full 72k corpus or "
        "the paper's 150-epoch/20-epoch protocols); denoise trained on DIV2K (BSD400/WED not found "
        "locally); frequency diagrams run at 256px (training/typical-inference scale), not at the "
        "larger resolutions TEST06 showed CAN produce a non-degenerate mask.",
    ]
    for i, line in enumerate(readme, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    ws = wb.create_sheet("Overall_PSNR_SSIM")
    write_df(ws, overall)

    ws = wb.create_sheet("Per_Degradation")
    pivot_psnr = by_deg.pivot(index="variant", columns="degradation", values="psnr").loc[VARIANTS].reset_index()
    pivot_ssim = by_deg.pivot(index="variant", columns="degradation", values="ssim").loc[VARIANTS].reset_index()
    write_df(ws, pivot_psnr)
    write_df(ws, pivot_ssim, start_row=len(pivot_psnr) + 4)

    ws = wb.create_sheet("Training_Loss")
    write_df(ws, final_loss)

    ws = wb.create_sheet("Per_Image_Raw")
    write_df(ws, per_image.head(2000))  # cap for workbook size

    ws = wb.create_sheet("Model_Config")
    write_df(ws, pd.DataFrame({
        "variant": VARIANTS,
        "mask_mode": ["none", "fixed(10x10)", "learned(MGB)", "learned(MGB)", "learned(MGB)"],
        "use_lh": [False, False, False, True, True],
        "use_hl": [False, False, False, False, True],
        "params": [26126644, 28717240, 28741600, 28765792, 28766086],
    }))

    ws = wb.create_sheet("Frequency_Diagrams")
    row_cursor = 1
    key_diagrams = ["C_learned_mask_fre1_Rain.png", "C_learned_mask_fre3_Haze.png",
                     "B_fixed_mask_fre3_Haze.png", "E_full_fre3_Haze.png"]
    for name in key_diagrams:
        p = DIAG_DIR / name
        if not p.exists():
            continue
        ws.cell(row=row_cursor, column=1, value=name).font = HEADER_FONT
        img = XLImage(str(p))
        img.width, img.height = 900, 190
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 13

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
