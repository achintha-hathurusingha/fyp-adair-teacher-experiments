"""TEST14: build the final 17-sheet Excel workbook LOCALLY. Run on the
local machine only, per the project's CSV-first / render-locally policy.

Usage (local Windows machine, from teacher-experiments/test14/scripts):
  python build_excel_local.py
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST14 = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST14 / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "AdaIR_Frequency_Augmented_Conditioning.xlsx"

HEADER_FONT = Font(bold=True)
BAND_COLS = [f"band{i+1}" for i in range(8)]


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def autosize_text_sheet(ws, width=115):
    ws.column_dimensions["A"].width = width


def main():
    q_df = pd.read_csv(STATS_DIR / "frequency_descriptors_all_crops.csv")
    validity_summary = pd.read_csv(STATS_DIR / "frequency_descriptor_validity_summary.csv")
    epoch_metrics = pd.read_csv(RESULTS_DIR / "epoch_metrics.csv")
    seed_summary = pd.read_csv(RESULTS_DIR / "seed_summary.csv")
    per_seed_deltas = pd.read_csv(STATS_DIR / "per_seed_deltas.csv")
    seed_level_stats = pd.read_csv(STATS_DIR / "seed_level_summary_stats.csv")
    per_deg_deltas = pd.read_csv(STATS_DIR / "per_degradation_deltas.csv")
    per_deg_summary = pd.read_csv(STATS_DIR / "per_degradation_summary.csv", header=[0, 1], index_col=[0, 1])
    control_df = pd.read_csv(STATS_DIR / "frequency_controls.csv")
    scene_var_df = pd.read_csv(STATS_DIR / "qF_scene_variance.csv")
    redundancy_probe = pd.read_csv(STATS_DIR / "freq_redundancy_probe.csv")
    redundancy_summary = pd.read_csv(STATS_DIR / "freq_redundancy_summary.csv")
    signature_summary = pd.read_csv(STATS_DIR / "freq_signature_summary.csv")
    coeff_df = pd.read_csv(STATS_DIR / "coefficient_analysis.csv")
    probe_df = pd.read_csv(STATS_DIR / "representation_probe.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----
    ws = wb.create_sheet("README")
    autosize_text_sheet(ws)
    t14f2 = seed_level_stats[(seed_level_stats.comparison == "T14-F2") &
                              (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    control_means = control_df[["psnr_normal", "psnr_zero_freq", "psnr_mean_freq", "psnr_shuffled_freq"]].mean()
    readme_lines = [
        "TEST14: Frequency-Augmented Degradation-Conditioned Operator",
        "",
        "Research question: after TEST06-R showed AdaIR's OWN frequency branch is causally "
        "irrelevant to its output, does an INDEPENDENT, compact frequency descriptor q_F "
        "(8-band radial FFT-magnitude profile, computed from scratch on the degraded input -- "
        "NOT reusing any AdaIR internal tensor) provide useful complementary information to "
        "TEST12's validated e_D + phi(F) conditional operator?",
        "",
        "PHASE 0 GATE (before training): q_F passed its validity audit -- 78.4% degradation-"
        "classification accuracy (well above the 33% chance level for 3 classes), non-trivial "
        "per-band variance, sensible signature (Haze concentrates 98.7% of spectral energy in "
        "the lowest band vs. 89-90% for Rain/Noise, consistent with haze being a smooth, "
        "low-frequency degradation).",
        "",
        f"HEADLINE FINDING: a clean, well-explained NO-GO / REDUNDANT FREQUENCY result. The 4 "
        f"mandatory causal controls are statistically identical -- normal={control_means['psnr_normal']:.3f}dB, "
        f"zero={control_means['psnr_zero_freq']:.3f}dB, mean={control_means['psnr_mean_freq']:.3f}dB, "
        f"shuffled={control_means['psnr_shuffled_freq']:.3f}dB (differences in the third decimal place, "
        "well within noise) -- meaning the trained coefficient generator learned to essentially ignore "
        "the actual content of q_F. The redundancy analysis explains why: adding q_F to "
        "[e_D, phi(F)] contributes ZERO additional degradation-classification accuracy (98.2% "
        "either way), canonical correlation between e_D and q_F is high (0.867), and "
        "between-degradation q_F distance (0.072) is nearly identical to within-scene "
        f"cross-degradation distance (0.075) -- q_F is driven by degradation identity, which "
        f"e_D (via the frozen teacher) already fully captures. T14-F2 restoration delta is "
        f"small and inconsistent (mean {t14f2['mean']:.3f}dB, same-sign only "
        f"{t14f2['same_sign_count']}/3 seeds).",
        "",
        "This does NOT claim AdaIR's frequency mechanism was correct, nor that frequency "
        "information is never useful -- it specifically means THIS descriptor, computed THIS "
        "way, adds no information beyond what the existing compact degradation embedding and "
        "spatial content summary already provide, for this student architecture and dataset.",
        "",
        "Directory isolation: reuses TEST07-B's dataset/KD-cache and TEST12's validated "
        "content-conditioned operator definition, READ-ONLY. q_F computed independently, "
        "explicitly NOT using any AdaIR internal FFT tensor. Imports fyp-adair-distill's "
        "locked NAFNet READ-ONLY. Does not modify any prior experiment directory.",
    ]
    for i, line in enumerate(readme_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    # ---- Frequency_Descriptor ----
    ws = wb.create_sheet("Frequency_Descriptor")
    desc_rows = pd.DataFrame({
        "field": ["formula", "bands", "normalization", "source", "dim"],
        "value": ["Y=0.299R+0.587G+0.114B -> |FFT2(Y)|^2 -> fftshift -> S/(sum(S)+eps) -> 8 radial bands",
                  "B1:0-1/16, B2:1/16-2/16, ..., B8:12/16-Nyquist (Nyquist=radius/(N/2)=1.0)",
                  "per-image, sums to ~1 (corner frequencies beyond axis-Nyquist excluded by design)",
                  "computed independently from the RGB degraded input -- explicitly NOT AdaIR's "
                  "raw_low/raw_high/M_l/M_h/FMiM/FMoM tensors", 8],
    })
    write_df(ws, desc_rows)
    start2 = len(desc_rows) + 3
    ws.cell(row=start2, column=1, value="Validity audit summary").font = HEADER_FONT
    write_df(ws, validity_summary, start_row=start2 + 1)

    # ---- Dataset ----
    ws = wb.create_sheet("Dataset")
    dataset_rows = pd.DataFrame({
        "field": ["source", "n_train_scenes", "n_val_scenes", "crops_per_train_scene", "crop_size_px",
                  "degradations"],
        "value": ["REUSED from test07_b/ (READ-ONLY)", 80, 20, 8, 128, "Rain, Haze, Noise"],
    })
    write_df(ws, dataset_rows)

    # ---- Models ----
    ws = wb.create_sheet("Models")
    models_df = pd.DataFrame({
        "model": ["A", "F2", "T14"],
        "description": ["Baseline locked NAFNet, no KD, no conditioning",
                         "= TEST12's validated fixed-basis operator: a=G([e_D;phi(F)])",
                         "F2 + q_F: a=G([e_D;phi(F);q_F]), same rank=2, same fixed basis"],
    })
    write_df(ws, models_df)

    # ---- Training_Config ----
    ws = wb.create_sheet("Training_Config")
    config_rows = pd.DataFrame({
        "field": ["epochs", "batch_size", "learning_rate", "optimizer", "lambda_kd", "seeds", "rank"],
        "value": [50, 8, "2e-4", "Adam", 0.1, "0, 1, 2", 2],
    })
    write_df(ws, config_rows)

    # ---- Epoch_Metrics ----
    ws = wb.create_sheet("Epoch_Metrics")
    write_df(ws, epoch_metrics)

    # ---- Seed_Summary ----
    ws = wb.create_sheet("Seed_Summary")
    write_df(ws, seed_summary)

    # ---- Restoration ----
    ws = wb.create_sheet("Restoration")
    write_df(ws, per_seed_deltas)
    start2 = len(per_seed_deltas) + 3
    ws.cell(row=start2, column=1, value="Seed-level summary (mean +- std, bootstrap 95% CI, same-sign count)").font = HEADER_FONT
    write_df(ws, seed_level_stats, start_row=start2 + 1)

    # ---- Per_Degradation ----
    ws = wb.create_sheet("Per_Degradation")
    write_df(ws, per_deg_deltas)
    start2 = len(per_deg_deltas) + 3
    ws.cell(row=start2, column=1, value="Per-degradation summary (mean +- std across 3 seeds)").font = HEADER_FONT
    per_deg_summary_flat = per_deg_summary.reset_index()
    per_deg_summary_flat.columns = ["comparison", "degradation", "delta_psnr_mean", "delta_psnr_std",
                                     "delta_ssim_mean", "delta_ssim_std"]
    write_df(ws, per_deg_summary_flat, start_row=start2 + 1)

    # ---- Frequency_Controls ----
    ws = wb.create_sheet("Frequency_Controls")
    write_df(ws, control_df)
    start2 = len(control_df) + 3
    ws.cell(row=start2, column=1, value="Mean PSNR by condition").font = HEADER_FONT
    cond_summary = pd.DataFrame({
        "condition": ["normal", "zero_freq", "mean_freq", "shuffled_freq"],
        "mean_psnr": [control_df.psnr_normal.mean(), control_df.psnr_zero_freq.mean(),
                      control_df.psnr_mean_freq.mean(), control_df.psnr_shuffled_freq.mean()],
    })
    write_df(ws, cond_summary, start_row=start2 + 1)

    # ---- Frequency_Distributions ----
    ws = wb.create_sheet("Frequency_Distributions")
    write_df(ws, q_df[["scene_id", "degradation", "split", "sum_qF"] + BAND_COLS])
    start2 = len(q_df) + 3
    ws.cell(row=start2, column=1, value="Scene variance by band/degradation").font = HEADER_FONT
    write_df(ws, scene_var_df, start_row=start2 + 1)
    start3 = start2 + len(scene_var_df) + 3
    ws.cell(row=start3, column=1, value="Frequency signature summary (between-deg vs within-scene-cross-deg)").font = HEADER_FONT
    write_df(ws, signature_summary, start_row=start3 + 1)

    # ---- Frequency_vs_Embedding ----
    ws = wb.create_sheet("Frequency_vs_Embedding")
    write_df(ws, redundancy_summary)
    start2 = len(redundancy_summary) + 3
    ws.cell(row=start2, column=1, value="Probe comparison (accuracy using different feature sets)").font = HEADER_FONT
    write_df(ws, redundancy_probe, start_row=start2 + 1)

    # ---- Frequency_vs_Content ----
    ws = wb.create_sheet("Frequency_vs_Content")
    ws.cell(row=1, column=1, value="See Frequency_vs_Embedding sheet -- redundancy_summary includes "
                                    "phi(F)_pca16 vs q_F correlation; redundancy_probe includes "
                                    "phi(F)_pca16-only and [phi(F)_pca16,q_F] rows for direct comparison.")

    # ---- Coefficient_Analysis ----
    ws = wb.create_sheet("Coefficient_Analysis")
    write_df(ws, coeff_df)

    # ---- Representation ----
    ws = wb.create_sheet("Representation")
    write_df(ws, probe_df)

    # ---- Complexity ----
    ws = wb.create_sheet("Complexity")
    a_row = seed_summary[seed_summary.model == "A"].iloc[0]
    f2_row = seed_summary[seed_summary.model == "F2"].iloc[0]
    t14_row = seed_summary[seed_summary.model == "T14"].iloc[0]
    complexity_rows = pd.DataFrame({
        "field": ["params_A", "params_F2", "params_T14", "extra_params_T14_vs_F2", "macs_A", "macs_F2",
                  "macs_T14", "extra_macs_T14_vs_F2", "note"],
        "value": [int(a_row.params), int(f2_row.params), int(t14_row.params),
                  int(t14_row.params) - int(f2_row.params), int(a_row.macs), int(f2_row.macs), int(t14_row.macs),
                  int(t14_row.macs) - int(f2_row.macs),
                  "NN parameter overhead is tiny (coefficient-head input expanded by 8 dims). The FFT "
                  "descriptor computation itself is a SEPARATE cost not captured by NN parameter count -- "
                  "this is an experimental information probe, NOT a claim of NPU-friendliness (FFT is not "
                  "NPU-friendly on Snapdragon Hexagon; a later experiment should test DCT/fixed cosine "
                  "filters/learned small filter banks as NPU-friendly replacements IF the signal were "
                  "useful -- it was not, so this is moot for TEST14)."],
    })
    write_df(ws, complexity_rows)

    # ---- GO_NO_GO ----
    ws = wb.create_sheet("GO_NO_GO")
    haze_delta = per_deg_summary.loc[("T14-F2", "haze"), ("delta_psnr", "mean")]
    rain_delta = per_deg_summary.loc[("T14-F2", "rain"), ("delta_psnr", "mean")]
    noise_delta = per_deg_summary.loc[("T14-F2", "noise"), ("delta_psnr", "mean")]
    shuffle_delta = (control_df.psnr_shuffled_freq - control_df.psnr_normal).mean()
    zero_delta = (control_df.psnr_zero_freq - control_df.psnr_normal).mean()
    go_rows = pd.DataFrame({
        "field": ["decision", "T14_minus_F2_mean_psnr", "T14_minus_F2_same_sign", "rain_delta", "haze_delta",
                  "noise_delta", "zero_freq_control_delta", "shuffled_freq_control_delta",
                  "probe_accuracy_with_qF", "probe_accuracy_without_qF", "cca_eD_qF", "rationale"],
        "value": ["NO-GO / REDUNDANT FREQUENCY", round(t14f2['mean'], 4) if False else None,
                  f"{t14f2['same_sign_count']}/3" if False else None,
                  round(rain_delta, 3), round(haze_delta, 3), round(noise_delta, 3),
                  round(zero_delta, 3), round(shuffle_delta, 3),
                  redundancy_probe[redundancy_probe.features == "[e_D,phi(F)_pca16,q_F]"].accuracy.iloc[0],
                  redundancy_probe[redundancy_probe.features == "[e_D,phi(F)_pca16]"].accuracy.iloc[0],
                  redundancy_summary[redundancy_summary.metric == "cca_top_correlation_eD_qF"].value.iloc[0],
                  "Two independent lines of evidence converge on the same conclusion. (1) Causal: all "
                  "4 frequency controls (normal/zero/mean/shuffled q_F) produce statistically identical "
                  "restoration PSNR -- the trained network ignores q_F's actual content. (2) "
                  "Informational: adding q_F to [e_D,phi(F)] yields ZERO additional degradation-probe "
                  "accuracy, and q_F correlates strongly with e_D (CCA=0.867) and is driven by "
                  "degradation identity (between-degradation distance approx. equals within-scene "
                  "cross-degradation distance). q_F itself is a real, non-trivial signal (78.4% probe "
                  "accuracy alone, passed the pre-training validity audit) -- but it is REDUNDANT with "
                  "information the teacher-distilled e_D embedding already provides. T14-F2 restoration "
                  "delta is small and inconsistent (2/3 same-sign), consistent with this redundancy. "
                  "Per the task's explicit rule, this branch should be closed."],
    })
    # fix the two None placeholders properly
    go_rows.loc[go_rows.field == "T14_minus_F2_mean_psnr", "value"] = round(t14f2['mean'], 4)
    go_rows.loc[go_rows.field == "T14_minus_F2_same_sign", "value"] = f"{int(t14f2['same_sign_count'])}/3"
    write_df(ws, go_rows)

    # ---- Environment ----
    ws = wb.create_sheet("Environment")
    env_rows = pd.DataFrame({
        "field": ["host", "gpu", "conda_env", "cpu_pinning", "concurrency"],
        "value": ["devon (192.248.10.68)", "RTX 4090", "adair-distill", "taskset -c 0-7,12-31",
                  "9-way concurrent (A/F2/T14 x 3 seeds)"],
    })
    write_df(ws, env_rows)

    # ---- Visualizations ----
    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 480, 300
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 18

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
