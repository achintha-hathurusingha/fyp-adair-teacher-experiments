"""TEST10-R: build the final 18-sheet Excel workbook LOCALLY. Run on the
local machine only, per the project's CSV-first / render-locally policy.

Usage (local Windows machine, from teacher-experiments/test10_r/src):
  python build_excel_local.py
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST10R = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST10R / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "AdaIR_Trajectory_Distillation_Corrected.xlsx"

HEADER_FONT = Font(bold=True)


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def autosize_text_sheet(ws, width=115):
    ws.column_dimensions["A"].width = width


def main():
    teacher_quality = pd.read_csv(RESULTS_DIR / "teacher_quality.csv")
    with open(RESULTS_DIR / "trajectory_targets_metadata.json") as f:
        traj_meta = json.load(f)
    target_collapse = pd.read_csv(STATS_DIR / "target_collapse_audit.csv")
    epoch_metrics = pd.read_csv(RESULTS_DIR / "epoch_metrics.csv")
    seed_summary = pd.read_csv(RESULTS_DIR / "seed_summary.csv")
    collapse_monitor = pd.read_csv(RESULTS_DIR / "collapse_monitor.csv")
    per_seed_deltas = pd.read_csv(STATS_DIR / "per_seed_deltas.csv")
    seed_level_stats = pd.read_csv(STATS_DIR / "seed_level_summary_stats.csv")
    per_deg_deltas = pd.read_csv(STATS_DIR / "per_degradation_deltas.csv")
    per_deg_summary = pd.read_csv(STATS_DIR / "per_degradation_summary.csv", header=[0, 1], index_col=[0, 1])
    probe_df = pd.read_csv(STATS_DIR / "representation_probe.csv")
    probe_agg = pd.read_csv(STATS_DIR / "representation_probe_aggregated.csv", header=[0, 1], index_col=0)
    stage_alignment = pd.read_csv(STATS_DIR / "stage_alignment.csv")
    cross_diversity = pd.read_csv(STATS_DIR / "cross_input_diversity.csv")
    residual_summary = pd.read_csv(STATS_DIR / "residual_analysis_summary.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----
    ws = wb.create_sheet("README")
    autosize_text_sheet(ws)
    readme_lines = [
        "TEST10-R: Corrected (Collapse-Resistant) Restoration-Trajectory Distillation",
        "",
        "TEST10 was INVALID: its jointly-trained teacher/student trajectory projection "
        "heads collapsed to near-constant vectors (cross-input pairwise cosine >0.999 for "
        "all 3 stages, all 3 seeds), so the near-zero training loss reflected a degenerate "
        "shortcut, not genuine trajectory alignment. TEST10-R fixes this by using FIXED, "
        "leakage-safe PCA-32 teacher-side targets (fit on training crops only, frozen, "
        "never updated by backprop) -- the same discipline already proven not to collapse "
        "for the established 16-dim KD embedding since TEST07-B.",
        "",
        "RESULT: this correction worked. The Phase-3 pre-training gate confirmed the fixed "
        "targets themselves are non-collapsed (mean pairwise cosine ~0, 97-98% degradation "
        "probe accuracy). Mandatory per-epoch collapse monitoring during all 9 training runs "
        "found ZERO collapse events (450/450 checks clean). Post-training stage alignment "
        "confirmed GENUINE, non-collapsed alignment: same-sample cosine 0.60-0.76 (real, "
        "substantial, but appropriately imperfect -- NOT the suspicious ~0.9999 seen in "
        "TEST10) while cross-input cosine stayed near 0 (matching the un-collapsed target's "
        "own baseline) in all 9/9 seed x stage combinations.",
        "",
        "Given this validity, the restoration result is now a TRUSTWORTHY test of H_TRAJ: "
        "Model G (trajectory distillation) underperforms Model F (TEST09's best mechanism) "
        "in ALL 3 seeds (mean delta -0.417dB, 95% CI entirely negative), and underperforms "
        "on ALL THREE degradations (Rain -0.77dB, Haze -0.41dB, Noise -0.08dB). This is a "
        "valid, informative NO-GO: once collapse is ruled out, transferring the teacher's "
        "intermediate restoration trajectory does not help this student, and mildly hurts it.",
        "",
        "A secondary correction: Phase 0's teacher-quality audit (using proper PSNR/SSIM, "
        "not TEST10's cruder raw-L2 residual) found the teacher is 'teacher_similar' "
        "(within +-0.5dB) rather than clearly worse on Haze -- refining TEST10's finding "
        "without changing its substance: Haze remains the one degradation where the teacher "
        "has no demonstrated restoration advantage, unlike its dramatic +5.7dB/+7.4dB leads "
        "on Rain/Noise.",
        "",
        "Directory isolation: reads test10/'s stage audit (reference only), test07_b's "
        "dataset/KD-cache, and teacher-experiments/scripts/instrument.py READ-ONLY. Imports "
        "fyp-adair-distill's locked NAFNet and the vendored AdaIR checkpoint READ-ONLY. Does "
        "not modify any prior experiment directory.",
    ]
    for i, line in enumerate(readme_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    # ---- Teacher_Quality ----
    ws = wb.create_sheet("Teacher_Quality")
    write_df(ws, teacher_quality)

    # ---- Teacher_Stages ----
    ws = wb.create_sheet("Teacher_Stages")
    stages_df = pd.DataFrame({
        "trajectory_stage": [0, 1, 2],
        "teacher_tensor": ["AFLB1.aflb_out", "AFLB2.aflb_out", "AFLB3.aflb_out"],
        "teacher_shape_128px": ["(1,384,16,16)", "(1,192,32,32)", "(1,96,64,64)"],
        "student_tensor": ["net.decoders[0] output", "net.decoders[1] output", "net.decoders[2] output"],
        "student_shape_128px": ["(1,128,16,16)", "(1,64,32,32)", "(1,32,64,64)"],
        "matched_by": ["spatial resolution (16x16)", "spatial resolution (32x32)", "spatial resolution (64x64)"],
        "source": ["Reused from test10/report/teacher_stage_audit.md (READ-ONLY, unchanged)"] * 3,
    })
    write_df(ws, stages_df)

    # ---- Target_PCA ----
    ws = wb.create_sheet("Target_PCA")
    target_pca_rows = pd.DataFrame({
        "stage": [0, 1, 2],
        "aflb_name": [traj_meta["stages"][str(i)]["aflb_name"] for i in range(3)],
        "raw_pooled_dim": [traj_meta["stages"][str(i)]["raw_pooled_dim"] for i in range(3)],
        "pca_dim": [32, 32, 32],
        "pca_fit_sample_count": [traj_meta["stages"][str(i)]["pca_fit_sample_count"] for i in range(3)],
        "pca_explained_variance_ratio_sum": [
            round(traj_meta["stages"][str(i)]["pca_explained_variance_ratio_sum"], 4) for i in range(3)],
        "fit_source": ["TRAINING crops ONLY (leakage-safe), frozen before student training"] * 3,
    })
    write_df(ws, target_pca_rows)

    # ---- Target_Collapse_Audit ----
    ws = wb.create_sheet("Target_Collapse_Audit")
    write_df(ws, target_collapse)

    # ---- Dataset ----
    ws = wb.create_sheet("Dataset")
    dataset_rows = pd.DataFrame({
        "field": ["source", "n_train_scenes", "n_val_scenes", "crops_per_train_scene", "crop_size_px",
                  "degradations"],
        "value": ["REUSED from test07_b/ (READ-ONLY, identical split/crops to TEST08-C/09/10)",
                  80, 20, 8, 128, "Rain, Haze, Noise"],
    })
    write_df(ws, dataset_rows)

    # ---- Training_Config ----
    ws = wb.create_sheet("Training_Config")
    config_rows = pd.DataFrame({
        "field": ["epochs", "batch_size", "learning_rate", "optimizer", "lambda_kd", "lambda_traj",
                  "stage_weights", "lowrank_rank", "traj_dim", "seeds", "student_stage_projection",
                  "teacher_stage_projection", "collapse_check_from_epoch", "collapse_cosine_threshold",
                  "loss_A", "loss_F", "loss_G"],
        "value": [50, 8, "2e-4", "Adam", 0.1, 0.1, "1/3 per stage (equal)", 2, 32, "0, 1, 2",
                  "Simple Linear(pooled_dim, 32) ONLY -- no MLP, no BatchNorm, trainable",
                  "FIXED StandardScaler+PCA-32 per stage, fit on training crops only, frozen, never "
                  "updated by backprop",
                  10, 0.98, "L_restore", "L_restore + 0.1*L_KD",
                  "L_restore + 0.1*L_KD + 0.1*L_traj (L_traj = mean of 3 stage MSE-after-L2norm losses "
                  "against FIXED targets)"],
    })
    write_df(ws, config_rows)

    # ---- Epoch_Metrics ----
    ws = wb.create_sheet("Epoch_Metrics")
    write_df(ws, epoch_metrics)

    # ---- Collapse_Monitor ----
    ws = wb.create_sheet("Collapse_Monitor")
    write_df(ws, collapse_monitor)

    # ---- Restoration (Seed_Summary + stats) ----
    ws = wb.create_sheet("Restoration")
    write_df(ws, seed_summary)
    start2 = len(seed_summary) + 3
    ws.cell(row=start2, column=1, value="Per-seed deltas").font = HEADER_FONT
    write_df(ws, per_seed_deltas, start_row=start2 + 1)
    start3 = start2 + len(per_seed_deltas) + 3
    ws.cell(row=start3, column=1, value="Seed-level summary (mean +- std, bootstrap 95% CI, N=3 EXPLORATORY)").font = HEADER_FONT
    write_df(ws, seed_level_stats, start_row=start3 + 1)

    # ---- Per_Degradation ----
    ws = wb.create_sheet("Per_Degradation")
    write_df(ws, per_deg_deltas)
    start2 = len(per_deg_deltas) + 3
    ws.cell(row=start2, column=1, value="Per-degradation summary (mean +- std across 3 seeds)").font = HEADER_FONT
    per_deg_summary_flat = per_deg_summary.reset_index()
    per_deg_summary_flat.columns = ["comparison", "degradation", "delta_psnr_mean", "delta_psnr_std",
                                     "delta_ssim_mean", "delta_ssim_std"]
    write_df(ws, per_deg_summary_flat, start_row=start2 + 1)

    # ---- Representation_Probe ----
    ws = wb.create_sheet("Representation_Probe")
    write_df(ws, probe_df)
    start2 = len(probe_df) + 3
    ws.cell(row=start2, column=1, value="Aggregated across seeds").font = HEADER_FONT
    probe_agg_flat = probe_agg.reset_index()
    probe_agg_flat.columns = ["representation", "accuracy_mean", "accuracy_std", "balanced_accuracy_mean",
                               "balanced_accuracy_std", "macro_f1_mean", "macro_f1_std"]
    write_df(ws, probe_agg_flat, start_row=start2 + 1)

    # ---- Stage_Alignment ----
    ws = wb.create_sheet("Stage_Alignment")
    write_df(ws, stage_alignment)

    # ---- Cross_Input_Diversity ----
    ws = wb.create_sheet("Cross_Input_Diversity")
    write_df(ws, cross_diversity)

    # ---- Residual_Analysis ----
    ws = wb.create_sheet("Residual_Analysis")
    write_df(ws, residual_summary)

    # ---- Complexity ----
    ws = wb.create_sheet("Complexity")
    a_row = seed_summary[seed_summary.model == "A"].iloc[0]
    f_row = seed_summary[seed_summary.model == "F"].iloc[0]
    g_row = seed_summary[seed_summary.model == "G"].iloc[0]
    complexity_rows = pd.DataFrame({
        "field": ["params_A", "params_F", "params_G_deployable", "params_G_training_incl_traj_heads",
                  "traj_heads_params_discarded_at_inference", "teacher_params_never_saved_to_student",
                  "macs_A", "macs_F", "macs_G", "note"],
        "value": [int(a_row.params), int(f_row.params), int(g_row.params) - 14432, int(g_row.params),
                  14432, 28784824, int(a_row.macs), int(f_row.macs), int(g_row.macs),
                  "verify_inference_graph.py confirms Model G's output and e_S are BIT-IDENTICAL with "
                  "or without traj_heads present -- deployable graph is self-contained, no AdaIR/PCA/"
                  "teacher-or-trajectory-projections required at inference. THEORETICAL COMPLEXITY "
                  "ONLY, not an NPU latency claim."],
    })
    write_df(ws, complexity_rows)

    # ---- Environment ----
    ws = wb.create_sheet("Environment")
    env_rows = pd.DataFrame({
        "field": ["host", "gpu", "conda_env", "cpu_pinning", "concurrency", "note"],
        "value": ["devon (192.248.10.68)", "RTX 4090", "adair-distill", "taskset -c 0-7,12-31",
                  "9-way concurrent (all A/F/G x 3 seeds together, ~2GB/process)",
                  "No online teacher needed this time (targets precomputed/fixed), unlike TEST10 -- "
                  "faster and lower memory than TEST10's approach"],
    })
    write_df(ws, env_rows)

    # ---- GO_NO_GO ----
    ws = wb.create_sheet("GO_NO_GO")
    gf_psnr = seed_level_stats[(seed_level_stats.comparison == "G-F") &
                                (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    haze_gf = per_deg_summary.loc[("G-F", "haze"), ("delta_psnr", "mean")]
    rain_gf = per_deg_summary.loc[("G-F", "rain"), ("delta_psnr", "mean")]
    noise_gf = per_deg_summary.loc[("G-F", "noise"), ("delta_psnr", "mean")]
    go_rows = pd.DataFrame({
        "field": ["decision", "mean_delta_GF_psnr_dB", "GF_all_3_seeds_same_sign", "rain_delta_GF_dB",
                  "haze_delta_GF_dB", "noise_delta_GF_dB", "any_training_collapse_detected",
                  "stage_alignment_valid_all_seeds_stages", "inference_graph_verified_clean",
                  "scientifically_valid", "rationale"],
        "value": ["NO-GO", round(gf_psnr["mean"], 3), "Yes, all 3 seeds negative", round(rain_gf, 3),
                  round(haze_gf, 3), round(noise_gf, 3), False, True, True, True,
                  "Unlike TEST10, this experiment IS scientifically valid: the fixed teacher targets "
                  "passed the pre-training collapse gate, zero collapse occurred during any of the 450 "
                  "epoch x stage monitoring checks across all 9 training runs, and post-training stage "
                  "alignment confirms genuine (same-sample cosine 0.60-0.76, non-degenerate) alignment "
                  "with healthy cross-input diversity (~0, matching the un-collapsed target baseline) "
                  "in all 9/9 seed x stage combinations. Given this validity, Model G (trajectory "
                  "distillation) consistently underperforms Model F (TEST09's best mechanism) -- "
                  "negative in all 3 seeds overall (-0.417dB) and on all three degradations (Rain "
                  "-0.77dB, Haze -0.41dB, Noise -0.08dB). This is a genuine NO-GO on restoration-"
                  "trajectory distillation as implemented here (3-stage, GAP+GMP+Linear projections, "
                  "MSE-after-L2norm loss, equal stage weights) -- NOT a methodology artifact this time. "
                  "Per the task's decision rule, do not add more trajectory stages; the correct next "
                  "step is a different mechanism, not more elaboration of this one."],
    })
    write_df(ws, go_rows)

    # ---- Visualizations ----
    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 480, 300
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 18

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
