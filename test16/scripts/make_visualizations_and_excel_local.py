"""TEST16: build visualizations and the final Excel workbook LOCALLY, from
the real Snapdragon 8 Elite QRD full-graph benchmark results pulled back
from devon. Run on the local machine only (CSV-first / render-locally
policy, established since TEST07-B).
"""
from __future__ import annotations

import json
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST16 = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST16 / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "Snapdragon_Full_Student_Benchmark.xlsx"
HEADER_FONT = Font(bold=True)
MODEL_NAMES = ["A", "F2", "N", "S"]
MODEL_COLORS = {"A": "#1f77b4", "F2": "#ff7f0e", "N": "#2ca02c", "S": "#d62728"}


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def main():
    pt_val = pd.read_csv(STATS_DIR / "pytorch_validation_summary.csv")
    full_graph = pd.read_csv(STATS_DIR / "full_graph_benchmark.csv")
    int8 = pd.read_csv(STATS_DIR / "int8_benchmark.csv")
    hotspots_top10 = pd.read_csv(STATS_DIR / "layer_hotspots_top10.csv")
    hotspots_bucket = pd.read_csv(STATS_DIR / "layer_hotspots_by_bucket.csv")
    with open(RESULTS_DIR / "export_manifest.json") as f:
        export_manifest = {e["name"]: e for e in json.load(f)}

    VIZ_DIR.mkdir(parents=True, exist_ok=True)

    # ---- master quality/hardware table (Phase 9) ----
    master_rows = []
    for name in MODEL_NAMES:
        v = pt_val[pt_val.model == name].iloc[0]
        g = full_graph[full_graph.name == name].iloc[0]
        e = export_manifest[name]
        master_rows.append({
            "Model": name, "Trained": bool(v.trained),
            "PSNR_dB": v.val_psnr, "SSIM": v.val_ssim,
            "Params": int(v.params), "MACs": int(v.macs),
            "Model_MB": round(e["model_size_bytes"] / 1e6, 2),
            "NPU_latency_mean_ms": g.latency_mean_ms, "NPU_latency_p95_ms": g.latency_p95_ms,
            "Peak_memory_MB": round(g.peak_memory_bytes / 1e6, 2),
            "CPU_fallback": bool(g.any_cpu_fallback), "NPU_only": bool(g.npu_only),
        })
    master_df = pd.DataFrame(master_rows)
    master_df.to_csv(STATS_DIR / "master_quality_hardware_table.csv", index=False)

    # ---- Viz 1: FP32 full-graph latency, log scale ----
    fig, ax = plt.subplots(figsize=(8, 5))
    colors = [MODEL_COLORS[n] for n in full_graph.name]
    bars = ax.bar(full_graph.name, full_graph.latency_mean_ms, color=colors,
                   yerr=full_graph.latency_std_ms, capsize=4)
    ax.set_yscale("log")
    ax.set_ylabel("Mean NPU latency (ms), log scale, n=100 warm reps")
    ax.set_title("TEST16: Full-Graph FP32 NPU Latency (Snapdragon 8 Elite QRD)")
    for bar, v in zip(bars, full_graph.latency_mean_ms):
        ax.text(bar.get_x() + bar.get_width() / 2, v * 1.15, f"{v:.0f}ms", ha="center")
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "01_fp32_full_graph_latency.png", dpi=150)
    plt.close(fig)

    # ---- Viz 2: INT8 latency ----
    fig, ax = plt.subplots(figsize=(8, 5))
    colors2 = [MODEL_COLORS[n] for n in int8.name]
    bars = ax.bar(int8.name, int8.latency_ms, color=colors2)
    ax.set_ylabel("INT8 NPU latency (ms)")
    ax.set_title("TEST16: Full-Graph INT8 NPU Latency")
    for bar, v in zip(bars, int8.latency_ms):
        ax.text(bar.get_x() + bar.get_width() / 2, v + 30, f"{v:.0f}ms", ha="center")
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "02_int8_full_graph_latency.png", dpi=150)
    plt.close(fig)

    # ---- Viz 3: FP32 vs INT8 side by side ----
    fig, ax = plt.subplots(figsize=(9, 5))
    x = range(len(MODEL_NAMES))
    fp32_vals = [full_graph[full_graph.name == n].iloc[0].latency_mean_ms for n in MODEL_NAMES]
    int8_vals = [int8[int8.name == n].iloc[0].latency_ms for n in MODEL_NAMES]
    width = 0.35
    ax.bar([i - width / 2 for i in x], fp32_vals, width, label="FP32", color="#1f77b4")
    ax.bar([i + width / 2 for i in x], int8_vals, width, label="INT8", color="#ff7f0e")
    ax.set_yscale("log")
    ax.set_xticks(list(x))
    ax.set_xticklabels(MODEL_NAMES)
    ax.set_ylabel("Latency (ms), log scale")
    ax.set_title("FP32 vs INT8: the Model-S FP32 anomaly does not survive quantization")
    ax.legend()
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "03_fp32_vs_int8.png", dpi=150)
    plt.close(fig)

    # ---- Viz 4: LayerNorm/AffineClamp cycle share by model ----
    ln_share = hotspots_bucket[hotspots_bucket.bucket == "LayerNorm/AffineClamp"]
    fig, ax = plt.subplots(figsize=(7, 5))
    ax.bar(ln_share.model, ln_share.pct_of_total_cycles, color=[MODEL_COLORS[n] for n in ln_share.model])
    ax.set_ylabel("% of total NPU execution cycles")
    ax.set_title("LayerNorm/AffineClamp Share of Full-Graph NPU Cycles")
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "04_layernorm_cycle_share.png", dpi=150)
    plt.close(fig)

    # ---- Viz 5: Pareto -- latency vs PSNR ----
    fig, ax = plt.subplots(figsize=(8, 6))
    for _, row in master_df.iterrows():
        marker = "o" if row.Trained else "x"
        ax.scatter(row.NPU_latency_mean_ms, row.PSNR_dB if pd.notna(row.PSNR_dB) else 0,
                   s=150, color=MODEL_COLORS[row.Model], marker=marker, label=row.Model)
        label = f"{row.Model}" + ("" if row.Trained else " (untrained,\nno PSNR axis)")
        ax.annotate(label, (row.NPU_latency_mean_ms, row.PSNR_dB if pd.notna(row.PSNR_dB) else 0),
                    textcoords="offset points", xytext=(8, 8))
    ax.set_xscale("log")
    ax.set_xlabel("Full-graph NPU latency (ms), log scale")
    ax.set_ylabel("Validation PSNR (dB) -- 0 = not measured (untrained)")
    ax.set_title("Quality/Latency Pareto (o=trained/real quality, x=untrained/latency-only)")
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "05_pareto_latency_psnr.png", dpi=150)
    plt.close(fig)

    # ---- Viz 6: latency vs peak memory ----
    fig, ax = plt.subplots(figsize=(8, 6))
    for _, row in master_df.iterrows():
        ax.scatter(row.NPU_latency_mean_ms, row.Peak_memory_MB, s=150, color=MODEL_COLORS[row.Model])
        ax.annotate(row.Model, (row.NPU_latency_mean_ms, row.Peak_memory_MB),
                    textcoords="offset points", xytext=(8, 8))
    ax.set_xscale("log")
    ax.set_xlabel("Full-graph NPU latency (ms), log scale")
    ax.set_ylabel("Peak inference memory (MB)")
    ax.set_title("Latency vs Peak Memory")
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "06_latency_vs_memory.png", dpi=150)
    plt.close(fig)

    print(f"wrote 6 visualizations to {VIZ_DIR}")

    # ================= Excel workbook =================
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 130
    readme = [
        "TEST16: Full Student Graph Snapdragon NPU Validation",
        "",
        "Objective: TEST15 benchmarked isolated micro-ops on the Snapdragon 8 Elite QRD "
        "(Hexagon V79) NPU. TEST16 compiles and profiles the ACTUAL complete restoration "
        "student graphs -- baseline (A), Minura's validated low-rank operator (F2), a "
        "normalization-surgery ablation (N), and a static-mixture reinterpretation (S) -- "
        "end to end, to see whether TEST15's per-op conclusions survive at full-graph scale.",
        "",
        "SCOPING NOTE: Models A and F2 have real trained checkpoints (TEST12). Models N and S "
        "do NOT -- per an explicit user decision this pass, they are exported/compiled/profiled "
        "UNTRAINED (architecturally real, numerically finite, but randomly initialized). This "
        "matches fyp-adair-distill's own established convention (src/models/norms.py): NPU "
        "latency does not depend on weight values, so untrained profiling is valid for latency "
        "and architecture claims, but PSNR/SSIM for N and S are NOT reported here (would be "
        "fabricated data) -- only A and F2 have real quality numbers this pass.",
        "",
        "HEADLINE FINDINGS:",
        "1. All 4 full graphs compiled and ran 100% on NPU in FP32 -- zero CPU/GPU fallback at "
        "full-graph scale, matching TEST15's isolated-op finding.",
        "2. Model N (LayerNorm2d -> affine_clamp, the ONLY change vs Model A, confirmed at the "
        "ONNX level: 0 ReduceMean/Sqrt/Pow vs A's 128/64/64) measures 430ms vs A's 10,528ms -- "
        "a ~24x latency reduction, and cycle-accurate (63.9M cycles for A vs 2.9M for N). This is "
        "a clean, architecture-isolated, high-confidence finding: normalization surgery is by far "
        "the single highest-leverage lever found in this project for NPU latency.",
        "3. Model S (static-mixture) measures 462ms in FP32 -- matching N, NOT A/F2 -- despite its "
        "ONNX graph containing the SAME LayerNorm2d ops as A/F2 (129 ReduceMean etc., confirmed). "
        "This is flagged as an UNRESOLVED finding: verified NOT a job-mixup (distinct model IDs, "
        "S-specific layer names present, stable 100-sample timings), but the compiler mechanism "
        "that appears to erase S's LayerNorm cost is not understood. Under INT8, this anomaly "
        "DISAPPEARS -- S reverts to A/F2-like cost (2530ms vs A's 2525ms), while N stays fastest "
        "(1609ms) in both precisions. This is read as evidence the FP32 S result is a compiler "
        "artifact specific to that precision/graph combination, not a generalizable property of "
        "the static-mixture architecture -- treat S's FP32 number with caution.",
        "4. INT8: real quantize+compile+profile+on-device-inference (not placeholder calibration) "
        "for all 4 models. A: 24.84dB (-2.44dB vs FP32's 27.28dB). F2: 26.01dB (-1.14dB vs FP32's "
        "27.15dB) -- F2 degrades LESS under INT8 than A does. All 4 remain 100% NPU with zero "
        "fallback under INT8 too.",
        "5. Model S's static-weight audit PASSED: every Conv node in its ONNX graph resolves to a "
        "compile-time-constant initializer -- confirmed programmatically, not by construction alone.",
        "",
        "Device: Snapdragon 8 Elite QRD (SM8750, Hexagon V79) -- identical device to TEST15. "
        "Runtime: QNN context binary. Input/output resolution: 256x256 (production export path, "
        "fyp-adair-distill/configs/export/qnn_int8.yaml), opset 17.",
        "",
        "Caveat: PyTorch-level validation (PSNR/SSIM for A/F2) used the TEST12 128x128 crop "
        "protocol; INT8 quality checks used the same validation images resized 128->256 to match "
        "the production export resolution (12-sample real on-device inference, not a full sweep) "
        "-- flagged as a modest-sample-size caveat, not a blocker.",
    ]
    for i, line in enumerate(readme, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    ws = wb.create_sheet("Environment")
    env = pd.DataFrame({
        "field": ["target_device", "chipset", "npu", "runtime", "onnx_opset", "input_resolution",
                  "qai_hub_version", "torch_version", "onnx_version", "precision_modes_tested"],
        "value": ["Snapdragon 8 Elite QRD", "SM8750", "Hexagon V79", "qnn_context_binary", 17,
                  "256x256", "0.53.0", "2.5.1+cu124", "1.17.0", "FP32, INT8"],
    })
    write_df(ws, env)

    ws = wb.create_sheet("Models")
    models_desc = pd.DataFrame({
        "Model": ["A", "F2", "N", "S"],
        "Description": [
            "Baseline NAFNet, no conditional mechanism. Trained (TEST12 checkpoint).",
            "Minura's validated rank-2 low-rank conditional operator (TEST08-C..TEST14 lineage). Trained (TEST12 checkpoint).",
            "Same as A but internal norm_type layernorm2d -> affine_clamp (fyp-adair-distill's own documented axis). UNTRAINED this pass.",
            "Static-mixture reinterpretation: 3 static 1x1-conv branches, runtime scalar mixing only. UNTRAINED this pass.",
        ],
        "Trained": [True, True, False, False],
    })
    write_df(ws, models_desc)

    ws = wb.create_sheet("PyTorch_Validation")
    write_df(ws, pt_val)

    ws = wb.create_sheet("ONNX_Export")
    export_rows = pd.DataFrame([{
        "name": n, "export_status": export_manifest[n]["export_status"],
        "node_count": export_manifest[n]["node_count"], "model_size_MB": round(export_manifest[n]["model_size_bytes"] / 1e6, 2),
        "n_conv": export_manifest[n]["n_conv"], "n_conv_dynamic_weight": export_manifest[n]["n_conv_dynamic_weight"],
        "all_conv_weights_constant": export_manifest[n]["all_conv_weights_constant"],
        "gather_count": export_manifest[n]["gather_count"], "fft_count": export_manifest[n]["fft_count"],
    } for n in MODEL_NAMES])
    write_df(ws, export_rows)

    ws = wb.create_sheet("Compilation")
    write_df(ws, full_graph[["name", "compile_success", "profile_success", "device", "compute_units_used",
                              "npu_only", "any_cpu_fallback", "any_gpu_fallback", "n_layers", "n_cpu_layers",
                              "n_gpu_layers"]])

    ws = wb.create_sheet("Full_Profile")
    write_df(ws, full_graph)

    ws = wb.create_sheet("Layer_Hotspots")
    write_df(ws, hotspots_top10)

    ws = wb.create_sheet("Normalization")
    norm_rows = master_df[master_df.Model.isin(["A", "N"])].copy()
    write_df(ws, norm_rows)
    r0 = len(norm_rows) + 3
    ws.cell(row=r0, column=1, value="Finding").font = HEADER_FONT
    ws.cell(row=r0 + 1, column=1,
            value="A (layernorm2d) vs N (affine_clamp), architecture otherwise identical: "
                  "10,528ms -> 430ms (~24.5x), 63.9M -> 2.9M cycles (~22x). ONNX-graph-confirmed: "
                  "N has 0 ReduceMean/Sqrt/Pow vs A's 128/64/64. This is the single highest-leverage "
                  "NPU latency lever identified across TEST15+TEST16.")

    ws = wb.create_sheet("Static_Mixture")
    sm_rows = master_df[master_df.Model.isin(["F2", "S"])].copy()
    write_df(ws, sm_rows)
    r0 = len(sm_rows) + 3
    ws.cell(row=r0, column=1, value="Finding (flagged unresolved)").font = HEADER_FONT
    ws.cell(row=r0 + 1, column=1,
            value="S's static-weight audit PASSED (every Conv weight is a compile-time constant). "
                  "FP32: S measures 462ms, matching N (430ms) despite S's ONNX graph containing the "
                  "same LayerNorm2d ops as A/F2 (129 ReduceMean, confirmed). Verified not a job-mixup "
                  "(distinct model IDs, S-specific layer names present in profile). Under INT8 this "
                  "anomaly disappears: S=2530ms, matching A(2525)/F2(2558), not N(1609). Read as: the "
                  "FP32 S speed is a compiler-specific artifact, not a confirmed architectural property "
                  "-- do not generalize from it without a dedicated follow-up.")

    ws = wb.create_sheet("INT8")
    write_df(ws, int8)

    ws = wb.create_sheet("Quality")
    write_df(ws, master_df[["Model", "Trained", "PSNR_dB", "SSIM"]])

    ws = wb.create_sheet("Latency")
    write_df(ws, master_df[["Model", "NPU_latency_mean_ms", "NPU_latency_p95_ms"]])

    ws = wb.create_sheet("Memory")
    write_df(ws, master_df[["Model", "Peak_memory_MB", "Model_MB"]])

    ws = wb.create_sheet("Pareto")
    write_df(ws, master_df)

    ws = wb.create_sheet("Operator_Map")
    write_df(ws, hotspots_bucket)

    ws = wb.create_sheet("Hardware_Constraints")
    constraints = pd.DataFrame({
        "id": ["C1", "C2", "C3", "C4", "C5"],
        "constraint": [
            "Static (compile-time-constant) Conv weights are cheap and universally NPU-native "
            "(TEST15: 54-139ms isolated; TEST16: A/F2/N/S all compile 100% NPU, zero fallback).",
            "Runtime-generated Conv weights are catastrophic (TEST15: dynamic_conv 4059ms, "
            "30-75x isolated-op penalty; not re-tested at full-graph scale in TEST16 -- none of "
            "A/F2/N/S use this pattern).",
            "Runtime SCALAR/channel mixing (Minura's low-rank op, static-mixture op) is cheap: "
            "both measured 74ms in TEST15 isolation, identical to a plain static conv.",
            "LayerNorm2d (mean/var/sqrt/div decomposition) is the dominant full-graph NPU cost: "
            "26% of A/F2's cycles directly, but its presence also blocks fusion of neighboring "
            "ops -- removing it (Model N) cuts total cycles ~22x, far more than 26% would predict.",
            "Static-mixture's full-graph interaction with LayerNorm-heavy backbones is UNCONFIRMED: "
            "Model S showed an unexplained FP32-only cost collapse that did not survive INT8 "
            "quantization -- do not treat S's FP32 number as a validated hardware-friendliness claim "
            "without a dedicated follow-up experiment.",
        ],
        "confidence": ["confirmed (TEST15+TEST16)", "confirmed (TEST15 isolated; untested at full-graph scale)",
                       "confirmed (TEST15 isolated)", "confirmed (TEST16, ONNX-graph-verified)",
                       "unresolved / flagged for follow-up"],
    })
    write_df(ws, constraints)

    ws = wb.create_sheet("GO_NO_GO")
    go_rows = pd.DataFrame({
        "question": [
            "Is Minura's low-rank operator (F2) NPU-safe?",
            "Does normalization surgery (N) preserve architecture while cutting latency?",
            "Is the static-mixture redesign (S) confirmed hardware-friendly?",
            "Should TEST17 retrain N and/or S to get real quality numbers?",
        ],
        "answer": [
            "YES -- F2 compiles 100% NPU, zero fallback, in both FP32 and INT8; quality is "
            "close to A (27.15 vs 27.28dB FP32; INT8 drop smaller than A's).",
            "YES, with high confidence -- the single largest latency lever found in this project "
            "(~24x FP32, ~1.6x INT8), architecture-isolated and cycle-verified.",
            "NOT YET -- FP32 result is real but unexplained and does not replicate under INT8. "
            "Needs a dedicated follow-up before being trusted as an architectural finding.",
            "YES -- N's latency case is now strong enough to justify the training cost; S needs "
            "the FP32 anomaly resolved first before quality numbers would be worth generating.",
        ],
    })
    write_df(ws, go_rows)

    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 560, 380
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 22

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
