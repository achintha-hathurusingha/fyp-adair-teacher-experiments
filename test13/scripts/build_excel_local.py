"""TEST13: build the final 16-sheet Excel workbook LOCALLY. Run on the
local machine only, per the project's CSV-first / render-locally policy.

Usage (local Windows machine, from teacher-experiments/test13/scripts):
  python build_excel_local.py
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST13 = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST13 / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "AdaIR_Adaptive_Basis_Conditioning.xlsx"

HEADER_FONT = Font(bold=True)


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def autosize_text_sheet(ws, width=115):
    ws.column_dimensions["A"].width = width


def main():
    epoch_metrics = pd.read_csv(RESULTS_DIR / "epoch_metrics.csv")
    seed_summary = pd.read_csv(RESULTS_DIR / "seed_summary.csv")
    per_seed_deltas = pd.read_csv(STATS_DIR / "per_seed_deltas.csv")
    seed_level_stats = pd.read_csv(STATS_DIR / "seed_level_summary_stats.csv")
    per_deg_deltas = pd.read_csv(STATS_DIR / "per_degradation_deltas.csv")
    per_deg_summary = pd.read_csv(STATS_DIR / "per_degradation_summary.csv", header=[0, 1], index_col=[0, 1])
    basis_df = pd.read_csv(STATS_DIR / "basis_adaptation.csv")
    rank_df = pd.read_csv(STATS_DIR / "basis_effective_rank.csv")
    control_df = pd.read_csv(STATS_DIR / "content_controls.csv")
    probe_df = pd.read_csv(STATS_DIR / "representation_probe.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----
    ws = wb.create_sheet("README")
    autosize_text_sheet(ws)
    t13f2 = seed_level_stats[(seed_level_stats.comparison == "T13-F2") &
                              (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    readme_lines = [
        "TEST13: Adaptive Low-Rank Operator Basis",
        "",
        "Research question: TEST12 established that the operator benefits from BOTH degradation "
        "state and current spatial content (shuffled-content control cost -2.91dB). TEST11 showed "
        "increasing coefficient rank does not help. H_BASIS: maybe the FIXED basis U0/V0 -- not the "
        "coefficients -- is the real constraint. TEST13 tests U(e_D)=U0+dU(e_D), V(e_D)=V0+dV(e_D), "
        "same rank=2, same coefficient generator as TEST12's validated operator (relabeled 'F2' here).",
        "",
        f"HEADLINE FINDING: NEGATIVE, and instructively so. T13 underperforms F2 in ALL 3 seeds for "
        f"both PSNR (mean {t13f2['mean']:.3f}dB, 95% CI [{t13f2['bootstrap_ci95_lo']:.3f}, "
        f"{t13f2['bootstrap_ci95_hi']:.3f}], same-sign 3/3) and SSIM (3/3 same-sign). Critically, the "
        "basis corrections are NOT small: relative_basis_change is 8.5-13.7x the original basis norm "
        "(NOT 'light adaptation' as the design intended, despite zero-initialization). The strong "
        "content-causal signature established in TEST12 (shuffled-content -2.91dB) is largely WASHED "
        "OUT here (shuffled-content only -0.13dB); instead, shuffled-basis-state (donor e_D) becomes "
        "the worst condition (-0.65dB) -- the operator's causal structure has shifted away from "
        "content-sensitivity toward an unstable dependence on the basis-adaptation input itself. This "
        "matches the task's pre-specified 'INTERESTING NEGATIVE' category: basis corrections became "
        "large, but restoration did NOT improve -- it got measurably worse -- indicating this "
        "particular adaptive-basis operator family is not expressive in a beneficial direction, "
        "rather than being merely under-adapted.",
        "",
        "Directory isolation: reuses TEST07-B's dataset/KD-cache and TEST12's validated "
        "content-conditioned operator definition, READ-ONLY. Imports fyp-adair-distill's locked "
        "NAFNet READ-ONLY. Does not modify any prior experiment directory.",
    ]
    for i, line in enumerate(readme_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    # ---- Models ----
    ws = wb.create_sheet("Models")
    models_df = pd.DataFrame({
        "model": ["A", "F2", "T13"],
        "description": ["Baseline locked NAFNet, no KD, no conditioning",
                         "= TEST12's validated T12 operator: fixed basis U0/V0, a=G([e_D;phi(F)])",
                         "Adaptive basis: U(e_D)=U0+dU(e_D), V(e_D)=V0+dV(e_D) (small Linear heads "
                         "from e_D only, zero-init), same coefficient generator as F2"],
    })
    write_df(ws, models_df)

    # ---- Dataset ----
    ws = wb.create_sheet("Dataset")
    dataset_rows = pd.DataFrame({
        "field": ["source", "n_train_scenes", "n_val_scenes", "crops_per_train_scene", "crop_size_px",
                  "degradations"],
        "value": ["REUSED from test07_b/ (READ-ONLY)", 80, 20, 8, 128, "Rain, Haze, Noise"],
    })
    write_df(ws, dataset_rows)

    # ---- Training_Config ----
    ws = wb.create_sheet("Training_Config")
    config_rows = pd.DataFrame({
        "field": ["epochs", "batch_size", "learning_rate", "optimizer", "lambda_kd", "seeds", "rank",
                  "basis_adaptation_heads", "basis_init"],
        "value": [50, 8, "2e-4", "Adam", 0.1, "0, 1, 2", 2,
                  "Linear(16, 256*2) for dU and dV, from e_D only",
                  "zero-init (dU=dV=0 at start, T13 begins exactly equivalent to F2)"],
    })
    write_df(ws, config_rows)

    # ---- Epoch_Metrics ----
    ws = wb.create_sheet("Epoch_Metrics")
    write_df(ws, epoch_metrics)

    # ---- Seed_Summary ----
    ws = wb.create_sheet("Seed_Summary")
    write_df(ws, seed_summary)

    # ---- Restoration ----
    ws = wb.create_sheet("Restoration")
    write_df(ws, per_seed_deltas)
    start2 = len(per_seed_deltas) + 3
    ws.cell(row=start2, column=1, value="Seed-level summary (mean +- std, bootstrap 95% CI, same-sign count)").font = HEADER_FONT
    write_df(ws, seed_level_stats, start_row=start2 + 1)

    # ---- Per_Degradation ----
    ws = wb.create_sheet("Per_Degradation")
    write_df(ws, per_deg_deltas)
    start2 = len(per_deg_deltas) + 3
    ws.cell(row=start2, column=1, value="Per-degradation summary (mean +- std across 3 seeds)").font = HEADER_FONT
    per_deg_summary_flat = per_deg_summary.reset_index()
    per_deg_summary_flat.columns = ["comparison", "degradation", "delta_psnr_mean", "delta_psnr_std",
                                     "delta_ssim_mean", "delta_ssim_std"]
    write_df(ws, per_deg_summary_flat, start_row=start2 + 1)

    # ---- Basis_Adaptation ----
    ws = wb.create_sheet("Basis_Adaptation")
    write_df(ws, basis_df)

    # ---- Basis_Effective_Rank ----
    ws = wb.create_sheet("Basis_Effective_Rank")
    write_df(ws, rank_df)

    # ---- Content_Controls ----
    ws = wb.create_sheet("Content_Controls")
    write_df(ws, control_df)
    start2 = len(control_df) + 3
    ws.cell(row=start2, column=1, value="Mean PSNR by condition (across all val crops, all seeds)").font = HEADER_FONT
    cond_summary = pd.DataFrame({
        "condition": ["normal", "zero_eD", "mean_content", "shuffled_content", "shuffled_basis_state"],
        "mean_psnr": [control_df.psnr_normal.mean(), control_df.psnr_zero_eD.mean(),
                      control_df.psnr_mean_content.mean(), control_df.psnr_shuffled_content.mean(),
                      control_df.psnr_shuffled_basis_state.mean()],
    })
    write_df(ws, cond_summary, start_row=start2 + 1)

    # ---- Representation ----
    ws = wb.create_sheet("Representation")
    write_df(ws, probe_df)

    # ---- Complexity ----
    ws = wb.create_sheet("Complexity")
    a_row = seed_summary[seed_summary.model == "A"].iloc[0]
    f2_row = seed_summary[seed_summary.model == "F2"].iloc[0]
    t13_row = seed_summary[seed_summary.model == "T13"].iloc[0]
    complexity_rows = pd.DataFrame({
        "field": ["params_A", "params_F2", "params_T13", "extra_params_T13_vs_F2", "pct_overhead_vs_A",
                  "macs_A", "macs_F2", "macs_T13", "note"],
        "value": [int(a_row.params), int(f2_row.params), int(t13_row.params),
                  int(t13_row.params) - int(f2_row.params),
                  f"{(int(t13_row.params) - int(f2_row.params)) / a_row.params * 100:.4f}%",
                  int(a_row.macs), int(f2_row.macs), int(t13_row.macs),
                  "Overhead well under the 0.5% target (0.236%). THEORETICAL COMPLEXITY ONLY, not "
                  "an NPU latency claim."],
    })
    write_df(ws, complexity_rows)

    # ---- Statistics ----
    ws = wb.create_sheet("Statistics")
    write_df(ws, seed_level_stats)

    # ---- GO_NO_GO ----
    ws = wb.create_sheet("GO_NO_GO")
    haze_t13f2 = per_deg_summary.loc[("T13-F2", "haze"), ("delta_psnr", "mean")]
    rain_t13f2 = per_deg_summary.loc[("T13-F2", "rain"), ("delta_psnr", "mean")]
    noise_t13f2 = per_deg_summary.loc[("T13-F2", "noise"), ("delta_psnr", "mean")]
    shuffle_c_delta = (control_df.psnr_shuffled_content - control_df.psnr_normal).mean()
    shuffle_b_delta = (control_df.psnr_shuffled_basis_state - control_df.psnr_normal).mean()
    go_rows = pd.DataFrame({
        "field": ["decision", "T13_minus_F2_mean_psnr", "T13_minus_F2_same_sign", "rain_delta", "haze_delta",
                  "noise_delta", "relative_basis_change_mean", "shuffled_content_delta_T13",
                  "shuffled_content_delta_T12_reference", "shuffled_basis_state_delta", "rationale"],
        "value": ["NO-GO (INTERESTING NEGATIVE)", round(t13f2["mean"], 4), f"{t13f2['same_sign_count']}/3",
                  round(rain_t13f2, 3), round(haze_t13f2, 3), round(noise_t13f2, 3),
                  round(basis_df[["relative_basis_change_U", "relative_basis_change_V"]].values.mean(), 2),
                  round(shuffle_c_delta, 3), -2.91, round(shuffle_b_delta, 3),
                  "T13 underperforms F2 consistently (3/3 seeds, both PSNR and SSIM) across ALL "
                  "degradations -- Rain hurt most (-1.25dB), Haze also negative (-0.24dB), Noise "
                  "roughly flat. This is NOT a case of insufficient adaptation: the basis corrections "
                  "are large (8.5-13.7x the original basis norm on average), not the intended light "
                  "near-zero adjustment. Representation quality (probe accuracy, teacher alignment) "
                  "is unchanged from F2, ruling out a representation-quality explanation. The "
                  "TEST12-established content-causal signature is substantially weakened (shuffled-"
                  "content cost -2.91dB in TEST12's T12 vs only -0.13dB here), while a NEW "
                  "instability appears: shuffled-basis-state (donor e_D) is now the single worst "
                  "condition (-0.65dB), meaning the operator has become more sensitive to WHICH "
                  "degradation-embedding generated the basis than to the actual image content it is "
                  "supposed to restore. This is the task's pre-specified 'INTERESTING NEGATIVE' "
                  "outcome: basis correction became large, but restoration did not improve -- the "
                  "adaptive-basis operator family (unconstrained additive U/V correction from e_D) "
                  "is not expressive in a beneficial direction. Per the task's explicit rule, this "
                  "motivates a DIFFERENT operator formulation, not more basis capacity."],
    })
    write_df(ws, go_rows)

    # ---- Environment ----
    ws = wb.create_sheet("Environment")
    env_rows = pd.DataFrame({
        "field": ["host", "gpu", "conda_env", "cpu_pinning", "concurrency"],
        "value": ["devon (192.248.10.68)", "RTX 4090", "adair-distill", "taskset -c 0-7,12-31",
                  "9-way concurrent (A/F2/T13 x 3 seeds)"],
    })
    write_df(ws, env_rows)

    # ---- Visualizations ----
    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 480, 300
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 18

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
