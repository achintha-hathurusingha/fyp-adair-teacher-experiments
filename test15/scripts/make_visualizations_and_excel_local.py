"""TEST15: build visualizations and the final Excel workbook LOCALLY, from
the real Snapdragon 8 Elite QRD (Hexagon V79) benchmark results pulled back
from devon. Run on the local machine only, per the project's CSV-first /
render-locally policy.

Usage (local Windows machine, from teacher-experiments/test15/scripts):
  python make_visualizations_and_excel_local.py
"""
from __future__ import annotations

from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST15 = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST15 / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "Snapdragon_NPU_Operator_Benchmark.xlsx"
HEADER_FONT = Font(bold=True)

BASE_OPS = ["conv3x3", "depthwise_conv", "conv1x1", "add", "multiply", "relu_clamp", "sigmoid",
            "softmax", "global_avg_pool", "global_max_pool", "resize", "concatenate",
            "elementwise_affine", "layernorm2d", "rmsnorm_like", "dynamic_conv", "fft"]
COMBO_OPS = ["combo_conv_add", "combo_conv_clamp", "combo_conv_mul", "combo_conv_add_clamp",
             "combo_dwconv_pointwise", "combo_1x1_add_clamp", "combo_gap_linear_mul"]
MINURA_OPS = ["minura_lowrank_op", "static_mixture_op"]


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def main():
    df = pd.read_csv(STATS_DIR / "npu_operator_benchmark.csv")
    VIZ_DIR.mkdir(parents=True, exist_ok=True)

    # ---- Viz 1: latency bar chart, log scale, sorted, dynamic_conv highlighted ----
    plot_df = df[df.profile_success == True].sort_values("latency_ms")  # noqa: E712
    fig, ax = plt.subplots(figsize=(11, 6))
    colors = ["#d62728" if n == "dynamic_conv" else ("#2ca02c" if n in MINURA_OPS else "#1f77b4")
              for n in plot_df.name]
    ax.barh(plot_df.name, plot_df.latency_ms, color=colors)
    ax.set_xscale("log")
    ax.set_xlabel("Latency (ms), log scale")
    ax.set_title("Snapdragon 8 Elite QRD (Hexagon V79): Operator Latency\n"
                  "(red = dynamic-weight conv, green = Minura candidate operators)")
    ax.grid(alpha=0.3, axis="x")
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "01_latency_by_operator.png", dpi=150)
    plt.close(fig)

    # ---- Viz 2: base ops only, linear scale (excl. dynamic_conv, fft) ----
    base_plot = df[df.name.isin(BASE_OPS) & (df.name != "dynamic_conv") & df.profile_success.fillna(False)]
    base_plot = base_plot.sort_values("latency_ms")
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.barh(base_plot.name, base_plot.latency_ms, color="#1f77b4")
    ax.set_xlabel("Latency (ms)")
    ax.set_title("Base Operator Latency (excluding dynamic_conv outlier)")
    ax.grid(alpha=0.3, axis="x")
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "02_base_op_latency_linear.png", dpi=150)
    plt.close(fig)

    # ---- Viz 3: combinations ----
    combo_plot = df[df.name.isin(COMBO_OPS)].sort_values("latency_ms")
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.barh(combo_plot.name, combo_plot.latency_ms, color="#9467bd")
    ax.set_xlabel("Latency (ms)")
    ax.set_title("Small-Combination Latency (fusion/memory-movement effects)")
    ax.grid(alpha=0.3, axis="x")
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "03_combination_latency.png", dpi=150)
    plt.close(fig)

    # ---- Viz 4: THE key comparison -- Minura vs static-mixture vs dynamic_conv ----
    key_df = df[df.name.isin(["dynamic_conv", "minura_lowrank_op", "static_mixture_op", "conv3x3"])]
    key_df = key_df.set_index("name").loc[["conv3x3", "minura_lowrank_op", "static_mixture_op", "dynamic_conv"]]
    fig, ax = plt.subplots(figsize=(8, 5))
    colors4 = ["#1f77b4", "#2ca02c", "#2ca02c", "#d62728"]
    bars = ax.bar(key_df.index, key_df.latency_ms, color=colors4)
    ax.set_yscale("log")
    ax.set_ylabel("Latency (ms), log scale")
    ax.set_title("The Key Comparison: Minura's Operator vs. the Dynamic-Conv Risk Case")
    for bar, v in zip(bars, key_df.latency_ms):
        ax.text(bar.get_x() + bar.get_width() / 2, v * 1.15, f"{v:.0f}ms", ha="center")
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "04_minura_vs_dynamic_conv.png", dpi=150)
    plt.close(fig)

    print(f"wrote 4 visualizations to {VIZ_DIR}")

    # ================= Excel workbook =================
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 120
    readme_lines = [
        "TEST15: Snapdragon NPU Operator & Graph Benchmark",
        "",
        "Objective: before continuing to invent restoration operators (TEST08-14), measure what "
        "the ACTUAL target NPU (Hexagon V79, in the Snapdragon 8 Elite / Snapdragon 8 Elite QRD "
        "reference device) can execute, and how fast, via real compile+profile jobs submitted to "
        "Qualcomm AI Hub's device farm -- not theoretical operator-support tables.",
        "",
        "HEADLINE FINDINGS:",
        "1. 24/26 operators compiled and ran successfully, 100% on NPU (zero CPU/GPU fallback "
        "observed at this isolated-operator granularity) -- more permissive than initially feared, "
        "though this does not guarantee zero fallback inside a full, larger student-network graph.",
        "2. FFT fails to even export to ONNX (opset 17 has no aten::fft_fft2 support) -- disqualified "
        "before reaching the NPU question at all. Confirms the project's existing avoidance of FFT.",
        "3. dynamic_conv (literal per-sample generated convolution weights) compiles and reports "
        "clean NPU execution, but runs at 4059ms -- 30-75x slower than every other operator tested "
        "(54-139ms range). This is the empirical confirmation of Qualcomm's documented warning about "
        "dynamic weights: the problem is not a hard compile failure, it is a severe, silent "
        "performance cliff.",
        "4. THE KEY RESULT: Minura's actual validated low-rank operator (U*diag(a(e_D))*V^T, fixed "
        "U/V, only a small coefficient vector generated at runtime) measures 74ms -- statistically "
        "identical to the proposed NPU-native static-mixture redesign (also 74ms), and squarely in "
        "the middle of the base-operator latency range. Minura's operator was NEVER the 'dynamic "
        "convolution' risk case -- it already behaves like a static-basis + dynamic-scalar-mixing "
        "operator, which is exactly the NPU-safe pattern the redesign hypothesis proposed. No "
        "operator-family redesign is needed on latency grounds; TEST09-14's mechanism search was "
        "not wasted effort chasing an NPU-infeasible operator.",
        "",
        "Device: Snapdragon 8 Elite QRD (Qualcomm reference design, chipset SM8750, Hexagon V79 NPU, "
        "matching the user-specified 'Hexagon V790' target). Runtime: QNN context binary "
        "(--target_runtime qnn_context_binary), forcing NPU-targeted compilation.",
        "",
        "Caveats: these are ISOLATED micro-benchmarks (single op or small combo), batch=1, fp32 (no "
        "INT8 quantization tested in this pass -- would require calibration data and is a natural "
        "follow-up). Real student-network latency depends on the full graph's fusion, memory "
        "movement, and scheduling, which this benchmark does not directly measure -- but it does "
        "conclusively rule in/out specific operator PATTERNS (especially: avoid literal dynamic "
        "conv weight generation; Minura's existing mechanism is fine).",
    ]
    for i, line in enumerate(readme_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    ws = wb.create_sheet("Operator_Benchmark_Table")
    write_df(ws, df)

    ws = wb.create_sheet("Base_Operators")
    write_df(ws, df[df.name.isin(BASE_OPS)])

    ws = wb.create_sheet("Combinations")
    write_df(ws, df[df.name.isin(COMBO_OPS)])

    ws = wb.create_sheet("Minura_Key_Comparison")
    key_rows = df[df.name.isin(["conv3x3", "minura_lowrank_op", "static_mixture_op", "dynamic_conv"])]
    write_df(ws, key_rows)
    start2 = len(key_rows) + 3
    ws.cell(row=start2, column=1, value="Interpretation").font = HEADER_FONT
    ws.cell(row=start2 + 1, column=1,
            value="minura_lowrank_op and static_mixture_op are statistically identical (74ms each), "
                  "both close to plain conv3x3 (89ms). dynamic_conv is 30-75x slower than any other "
                  "operator tested. Minura's existing mechanism does not need to be redesigned around "
                  "the static-mixture pattern for latency reasons -- it already achieves that pattern's "
                  "performance.")

    ws = wb.create_sheet("Environment")
    env_rows = pd.DataFrame({
        "field": ["target_device", "chipset", "npu", "runtime", "onnx_opset", "batch_size", "precision",
                  "quantization_tested"],
        "value": ["Snapdragon 8 Elite QRD (Qualcomm reference design)", "SM8750", "Hexagon V79",
                  "qnn_context_binary (QNN/QAIRT)", 17, 1, "fp32",
                  "No (natural follow-up: INT8 with calibration data)"],
    })
    write_df(ws, env_rows)

    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 520, 330
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 20

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
