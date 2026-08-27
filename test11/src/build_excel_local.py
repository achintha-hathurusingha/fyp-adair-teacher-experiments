"""TEST11: build the final 17-sheet Excel workbook LOCALLY. Run on the
local machine only, per the project's CSV-first / render-locally policy.

Usage (local Windows machine, from teacher-experiments/test11/src):
  python build_excel_local.py
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST11 = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST11 / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "AdaIR_LowRank_Conditional_Operator.xlsx"

HEADER_FONT = Font(bold=True)
RANK_MODELS = ["F2", "F4", "F8", "F16"]
RANKS = {"F2": 2, "F4": 4, "F8": 8, "F16": 16}


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def autosize_text_sheet(ws, width=115):
    ws.column_dimensions["A"].width = width


def main():
    epoch_metrics = pd.read_csv(RESULTS_DIR / "epoch_metrics.csv")
    seed_summary = pd.read_csv(RESULTS_DIR / "seed_summary.csv")
    per_seed_deltas = pd.read_csv(STATS_DIR / "per_seed_deltas.csv")
    seed_level_stats = pd.read_csv(STATS_DIR / "seed_level_summary_stats.csv")
    per_deg_deltas = pd.read_csv(STATS_DIR / "per_degradation_deltas.csv")
    per_deg_summary = pd.read_csv(STATS_DIR / "per_degradation_summary.csv", header=[0, 1], index_col=[0, 1])
    coeff_df = pd.read_csv(STATS_DIR / "coefficient_analysis.csv")
    mod_df = pd.read_csv(STATS_DIR / "modulation_magnitude.csv")
    eff_rank_df = pd.read_csv(STATS_DIR / "effective_rank.csv")
    probe_df = pd.read_csv(STATS_DIR / "representation_probe.csv")
    align_df = pd.read_csv(STATS_DIR / "teacher_alignment.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----
    ws = wb.create_sheet("README")
    autosize_text_sheet(ws)
    readme_lines = [
        "TEST11: Low-Rank Conditional Operator Capacity",
        "",
        "Research question: TEST09 found rank-2 low-rank channel-mixing (Model F) to be the "
        "strongest conditioning mechanism so far, but it still underperforms baseline NAFNet "
        "overall and never rescues Haze. TEST11 asks whether increasing the rank R (2/4/8/16) "
        "closes that gap, testing the hypothesis that different degradations need different "
        "operator capacity -- specifically that Haze might need higher rank than Rain/Noise.",
        "",
        "HEADLINE FINDING: rank is not the bottleneck. All three rank-scaling steps (F4-F2, "
        "F8-F4, F16-F8) are inconsistent across seeds (same-sign count only 2/3 for each), "
        "essentially flat/noisy. Every rank (F2/F4/F8/F16) underperforms baseline A "
        "consistently (3/3 seeds negative each). Haze shows almost IDENTICAL underperformance "
        "at every rank (-0.84 to -1.09dB vs baseline, no trend). Critically, EFFECTIVE RANK "
        "(participation-ratio of the learned coefficient covariance) never exceeds ~2.6 "
        "regardless of configured rank -- F16's configured 16 dimensions are used at only "
        "~16% utilization (effective rank 2.63), barely more than F2's ~54% utilization of "
        "its much smaller budget (effective rank 1.08). The model consistently collapses "
        "toward using only 1-3 effective directions no matter how many are made available.",
        "",
        "This matches the task's own 'IMPORTANT FAILURE CASE': configured capacity is not "
        "being used, so increasing it further will not help -- the correct next step per the "
        "task's decision rule is to investigate COEFFICIENT GENERATION (the head that maps "
        "e_S to the coefficient vector a), not to increase rank again.",
        "",
        "Directory isolation: reuses TEST07-B's dataset/KD-cache and TEST09's compact-latent "
        "+ low-rank mechanism definition, READ-ONLY. Imports fyp-adair-distill's locked "
        "NAFNet READ-ONLY. Does not modify any prior experiment directory.",
    ]
    for i, line in enumerate(readme_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    # ---- Models ----
    ws = wb.create_sheet("Models")
    models_df = pd.DataFrame({
        "model": ["A", "F2", "F4", "F8", "F16"],
        "description": ["Baseline locked NAFNet, no KD, no conditioning",
                         "Compact latent KD + rank-2 low-rank channel mixing (= TEST09's Model F)",
                         "Same as F2, rank=4", "Same as F2, rank=8", "Same as F2, rank=16"],
        "mathematical_form": ["N/A"] + ["F' = F + U diag(a(e_S)) V^T F, U,V in R^(256xR), a(e_S) in R^R"] * 4,
    })
    write_df(ws, models_df)

    # ---- Dataset ----
    ws = wb.create_sheet("Dataset")
    dataset_rows = pd.DataFrame({
        "field": ["source", "n_train_scenes", "n_val_scenes", "crops_per_train_scene", "crop_size_px",
                  "degradations"],
        "value": ["REUSED from test07_b/ (READ-ONLY, identical split/crops to TEST08-C/09/10/10-R)",
                  80, 20, 8, 128, "Rain, Haze, Noise"],
    })
    write_df(ws, dataset_rows)

    # ---- Training_Config ----
    ws = wb.create_sheet("Training_Config")
    config_rows = pd.DataFrame({
        "field": ["epochs", "batch_size", "learning_rate", "optimizer", "lambda_kd", "seeds",
                  "ranks_tested", "rank_specific_LR_tuning", "rank_specific_KD_tuning"],
        "value": [50, 8, "2e-4", "Adam", 0.1, "0, 1, 2", "2, 4, 8, 16",
                  "None -- identical LR for all ranks per spec", "None -- identical lambda_kd=0.1 for all ranks"],
    })
    write_df(ws, config_rows)

    # ---- Epoch_Metrics ----
    ws = wb.create_sheet("Epoch_Metrics")
    write_df(ws, epoch_metrics)

    # ---- Seed_Summary ----
    ws = wb.create_sheet("Seed_Summary")
    write_df(ws, seed_summary)

    # ---- Restoration ----
    ws = wb.create_sheet("Restoration")
    write_df(ws, per_seed_deltas)
    start2 = len(per_seed_deltas) + 3
    ws.cell(row=start2, column=1, value="Seed-level summary (mean +- std, bootstrap 95% CI, same-sign count)").font = HEADER_FONT
    write_df(ws, seed_level_stats, start_row=start2 + 1)

    # ---- Per_Degradation ----
    ws = wb.create_sheet("Per_Degradation")
    write_df(ws, per_deg_deltas)
    start2 = len(per_deg_deltas) + 3
    ws.cell(row=start2, column=1, value="Per-degradation summary (mean +- std across 3 seeds)").font = HEADER_FONT
    per_deg_summary_flat = per_deg_summary.reset_index()
    per_deg_summary_flat.columns = ["comparison", "degradation", "delta_psnr_mean", "delta_psnr_std",
                                     "delta_ssim_mean", "delta_ssim_std"]
    write_df(ws, per_deg_summary_flat, start_row=start2 + 1)

    # ---- Rank_Comparison ----
    ws = wb.create_sheet("Rank_Comparison")
    degs = ["rain", "haze", "noise"]
    rank_table_rows = []
    for deg in degs:
        row = {"degradation": deg.capitalize()}
        for m in RANK_MODELS:
            row[f"R={RANKS[m]}"] = round(seed_summary[seed_summary.model == m][f"last5_mean_{deg}_psnr"].mean(), 3)
        rank_table_rows.append(row)
    rank_table_rows.append({"degradation": "Overall", **{
        f"R={RANKS[m]}": round(seed_summary[seed_summary.model == m].last5_mean_psnr.mean(), 3) for m in RANK_MODELS}})
    write_df(ws, pd.DataFrame(rank_table_rows))

    # ---- Coefficient_Analysis ----
    ws = wb.create_sheet("Coefficient_Analysis")
    write_df(ws, coeff_df)
    start2 = len(coeff_df) + 3
    ws.cell(row=start2, column=1, value="Modulation magnitude (raw per-sample rows)").font = HEADER_FONT
    write_df(ws, mod_df, start_row=start2 + 1)

    # ---- Effective_Rank ----
    ws = wb.create_sheet("Effective_Rank")
    write_df(ws, eff_rank_df)
    start2 = len(eff_rank_df) + 3
    ws.cell(row=start2, column=1, value="Effective rank vs configured rank, ALL degradations (mean across seeds)").font = HEADER_FONT
    eff_summary = eff_rank_df[eff_rank_df.degradation == "ALL"].groupby("model")[
        ["configured_rank", "effective_rank"]].mean().reset_index()
    eff_summary["utilization_pct"] = (eff_summary.effective_rank / eff_summary.configured_rank * 100).round(1)
    write_df(ws, eff_summary, start_row=start2 + 1)

    # ---- Representation_Probe ----
    ws = wb.create_sheet("Representation_Probe")
    write_df(ws, probe_df)

    # ---- Teacher_Alignment ----
    ws = wb.create_sheet("Teacher_Alignment")
    write_df(ws, align_df)

    # ---- Complexity ----
    ws = wb.create_sheet("Complexity")
    a_row = seed_summary[seed_summary.model == "A"].iloc[0]
    complexity_rows = []
    for m in RANK_MODELS:
        row = seed_summary[seed_summary.model == m].iloc[0]
        psnr_gain = row.last5_mean_psnr - a_row.last5_mean_psnr
        extra_params = int(row.params) - int(a_row.params)
        complexity_rows.append({
            "model": m, "rank": RANKS[m], "params": int(row.params), "extra_params_vs_A": extra_params,
            "macs": int(row.macs), "extra_macs_vs_A": int(row.macs) - int(a_row.macs),
            "delta_psnr_vs_A": round(psnr_gain, 4),
            "psnr_gain_per_1000_extra_params": round(psnr_gain / (extra_params / 1000), 5) if extra_params else None,
        })
    complexity_df = pd.DataFrame(complexity_rows)
    write_df(ws, complexity_df)
    start2 = len(complexity_df) + 3
    ws.cell(row=start2, column=1, value="Note: THEORETICAL COMPLEXITY ONLY, not an NPU latency claim.").font = HEADER_FONT

    # ---- Statistics ----
    ws = wb.create_sheet("Statistics")
    write_df(ws, seed_level_stats)

    # ---- GO_NO_GO ----
    ws = wb.create_sheet("GO_NO_GO")
    f4f2 = seed_level_stats[(seed_level_stats.comparison == "F4-F2") & (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    f8f4 = seed_level_stats[(seed_level_stats.comparison == "F8-F4") & (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    f16f8 = seed_level_stats[(seed_level_stats.comparison == "F16-F8") & (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    haze_f2a = per_deg_summary.loc[("F2-A", "haze"), ("delta_psnr", "mean")]
    haze_f16a = per_deg_summary.loc[("F16-A", "haze"), ("delta_psnr", "mean")]
    go_rows = pd.DataFrame({
        "field": ["decision", "F4_minus_F2_mean_dB", "F4_minus_F2_same_sign", "F8_minus_F4_mean_dB",
                  "F8_minus_F4_same_sign", "F16_minus_F8_mean_dB", "F16_minus_F8_same_sign",
                  "haze_delta_vs_A_at_R2", "haze_delta_vs_A_at_R16", "haze_improves_with_rank",
                  "max_effective_rank_observed", "effective_rank_saturates", "rationale"],
        "value": ["NO-GO", round(f4f2["mean"], 4), f"{f4f2['same_sign_count']}/3", round(f8f4["mean"], 4),
                  f"{f8f4['same_sign_count']}/3", round(f16f8["mean"], 4), f"{f16f8['same_sign_count']}/3",
                  round(haze_f2a, 3), round(haze_f16a, 3), False, 2.63, True,
                  "F2≈F4≈F8≈F16: none of the three rank-scaling steps show a consistent "
                  "(3/3 same-sign) restoration improvement, and Haze's deficit vs baseline is "
                  "essentially unchanged from R=2 (-0.89dB) to R=16 (-1.09dB) -- rank does not "
                  "rescue Haze. Root cause identified: effective rank (participation ratio of "
                  "the learned coefficient covariance) saturates at ~2.6 regardless of "
                  "configured rank up to 16 -- the extra configured capacity is simply not "
                  "being used by the coefficient-generation head. Representation quality "
                  "(probe accuracy, teacher-embedding cosine similarity) is also unchanged "
                  "across ranks, confirming the bottleneck is specifically in how the "
                  "low-rank operator's coefficients are generated/used, not in the compact "
                  "degradation representation itself. Per the task's explicit decision rule: "
                  "do NOT continue increasing rank; investigate coefficient generation instead."],
    })
    write_df(ws, go_rows)

    # ---- Environment ----
    ws = wb.create_sheet("Environment")
    env_rows = pd.DataFrame({
        "field": ["host", "gpu", "conda_env", "cpu_pinning", "wave1", "wave2"],
        "value": ["devon (192.248.10.68)", "RTX 4090", "adair-distill", "taskset -c 0-7,12-31",
                  "A/F2/F4 x 3 seeds (9-way concurrent)", "F8/F16 x 3 seeds (6-way concurrent)"],
    })
    write_df(ws, env_rows)

    # ---- Visualizations ----
    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 480, 300
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 18

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
