"""TEST08-C: build the final 16-sheet Excel workbook LOCALLY (Windows
machine), reading the CSVs pulled back from devon. Run on the local machine
only, per the project's CSV-first / render-locally policy.

Usage (local Windows machine, from teacher-experiments/test08_c/src):
  python build_excel_local.py
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST08C = Path(__file__).resolve().parent.parent
TEACHER_EXP = TEST08C.parent
RESULTS_DIR = TEST08C / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
TEST07B_TEACHER_META = TEACHER_EXP / "test07_b" / "results" / "teacher_cache" / "teacher_cache_metadata.json"
OUT_XLSX = RESULTS_DIR / "AdaIR_Degradation_Conditioned_Spatial.xlsx"

HEADER_FONT = Font(bold=True)


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def autosize_text_sheet(ws, width=110):
    ws.column_dimensions["A"].width = width


def main():
    with open(TEST07B_TEACHER_META) as f:
        teacher_meta = json.load(f)

    epoch_metrics = pd.read_csv(RESULTS_DIR / "epoch_metrics.csv")
    seed_summary = pd.read_csv(RESULTS_DIR / "seed_summary.csv")
    per_seed_deltas = pd.read_csv(STATS_DIR / "per_seed_deltas.csv")
    seed_level_stats = pd.read_csv(STATS_DIR / "seed_level_summary_stats.csv")
    per_deg_deltas = pd.read_csv(STATS_DIR / "per_degradation_deltas.csv")
    per_deg_summary = pd.read_csv(STATS_DIR / "per_degradation_summary.csv", header=[0, 1], index_col=[0, 1])
    probe_df = pd.read_csv(STATS_DIR / "representation_probe.csv")
    probe_agg = pd.read_csv(STATS_DIR / "representation_probe_aggregated.csv", header=[0, 1], index_col=0)
    align_df = pd.read_csv(STATS_DIR / "teacher_student_alignment.csv")
    cond_df = pd.read_csv(STATS_DIR / "conditioning_statistics.csv")
    change_df = pd.read_csv(STATS_DIR / "bottleneck_change.csv")
    control_df = pd.read_csv(STATS_DIR / "random_control.csv")
    interv_df = pd.read_csv(STATS_DIR / "embedding_intervention.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----
    ws = wb.create_sheet("README")
    autosize_text_sheet(ws, 115)
    readme_lines = [
        "TEST08-C: Compact Degradation State + Spatial Conditioning",
        "",
        "Research question: TEST07-B showed the student CAN match the teacher's compact "
        "degradation representation (cosine ~0.99, probe accuracy ~96%), but simple latent-MSE "
        "distillation did NOT improve restoration (mean delta -0.79dB vs baseline, Rain/Haze "
        "hurt, Noise helped). TEST08-C asks: does actively using that representation to "
        "CONDITION the spatial restoration computation (FiLM-style bottleneck modulation) "
        "convert it into a real restoration benefit?",
        "",
        "Directory isolation: reads test07_b/'s dataset manifest and teacher embedding cache "
        "READ-ONLY (verified identical split/crops, no regeneration needed), and "
        "fyp-adair-distill's locked NAFNet architecture READ-ONLY. Does not modify test07_b/, "
        "test07_pilot/, test06_r/, test06/, or earlier experiments. All outputs live under "
        "test08_c/.",
        "",
        "Models: A (baseline NAFNet), B (validated TEST07-B compact latent KD, reference), "
        "C (B + zero-initialized FiLM-style bottleneck conditioning from the SAME e_S used "
        "for the KD loss). Model D (oracle teacher-conditioned control) was NOT RUN, per the "
        "task's explicit instruction not to train it.",
        "",
        "Data-integrity note: uses the exact per-run-CSV training pattern established after "
        "TEST07-B's shared-file race-condition bug -- all 9 runs (A/B/C x 3 seeds) write their "
        "own output files, merged in a separate deterministic step. Model A and B numbers here "
        "exactly reproduce TEST07-B's own results (fully deterministic given identical seed, "
        "data, and architecture) -- a strong internal consistency check.",
        "",
        "Sheet guide:",
        "  Dataset                 - reused TEST07-B dataset (reference only, not regenerated)",
        "  Teacher                 - AdaIR checkpoint verification, reused PCA-16 details",
        "  Training_Config          - hyperparameters, identical-except-KD/conditioning confirmation",
        "  Epoch_Metrics            - full per-epoch training/validation history, all 9 runs",
        "  Seed_Summary             - per-run checkpoint-selection summary (best/final/last5)",
        "  Restoration              - primary statistical analysis: C-A, C-B, B-A paired deltas",
        "  Per_Degradation           - Rain/Haze/Noise breakdown for all three comparisons",
        "  Representation_Probe     - degradation-classification probe, Teacher/A/B/C",
        "  Conditioning_Statistics  - gamma/beta distribution by degradation (Model C)",
        "  Bottleneck_Change         - relative bottleneck modulation magnitude by degradation",
        "  Random_Control            - learned vs random/shuffled/zero conditioning comparison",
        "  Embedding_Intervention    - donor/recipient degradation-embedding swap causal test",
        "  Complexity                - parameter/MACs overhead, A vs B vs C",
        "  GO_NO_GO                  - final decision and rationale",
    ]
    for i, line in enumerate(readme_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    # ---- Dataset ----
    ws = wb.create_sheet("Dataset")
    dataset_rows = pd.DataFrame({
        "field": ["source", "n_train_scenes", "n_val_scenes", "n_crops_per_train_scene",
                  "n_crops_per_val_scene", "crop_size_px", "degradations", "patch_sampling_option"],
        "value": ["REUSED from test07_b/ (READ-ONLY, verified identical split/crops -- not regenerated)",
                  80, 20, 8, 1, 128, "Rain, Haze, Noise",
                  "OPTION 2: precomputed multi-crop cache (reused from test07_b)"],
    })
    write_df(ws, dataset_rows)

    # ---- Teacher ----
    ws = wb.create_sheet("Teacher")
    teacher_rows = pd.DataFrame({
        "field": ["checkpoint_path", "checkpoint_sha256", "sha256_matches_expected", "n_params", "frozen",
                  "raw_pooled_dim", "pca_dim", "pca_fit_sample_count", "pca_explained_variance_ratio_sum",
                  "reuse_source"],
        "value": [teacher_meta["teacher_checkpoint_path"], teacher_meta["teacher_checkpoint_sha256"],
                  teacher_meta["sha256_matches"], teacher_meta["n_params"], teacher_meta["teacher_frozen"],
                  teacher_meta["raw_pooled_dim"], teacher_meta["pca_dim"], teacher_meta["pca_fit_sample_count"],
                  round(teacher_meta["pca_explained_variance_ratio_sum"], 4),
                  "REUSED from test07_b/results/teacher_cache/ (READ-ONLY; identical training split/crops)"],
    })
    write_df(ws, teacher_rows)

    # ---- Training_Config ----
    ws = wb.create_sheet("Training_Config")
    config_rows = pd.DataFrame({
        "field": ["epochs", "batch_size", "learning_rate", "optimizer", "lambda_kd", "seeds", "models",
                  "identical_config_ABC", "conditioning_location", "conditioning_init",
                  "same_eS_for_KD_and_conditioning"],
        "value": [50, 8, "2e-4", "Adam", 0.1, "0, 1, 2", "A (baseline), B (compact latent KD), "
                  "C (KD + bottleneck FiLM conditioning)", True,
                  "deepest bottleneck ONLY (no encoder/skip/decoder conditioning)",
                  "zero-init gamma_head/beta_head -> gamma=1, beta=0 (identity) at init",
                  True],
    })
    write_df(ws, config_rows)

    # ---- Epoch_Metrics ----
    ws = wb.create_sheet("Epoch_Metrics")
    write_df(ws, epoch_metrics)

    # ---- Seed_Summary ----
    ws = wb.create_sheet("Seed_Summary")
    write_df(ws, seed_summary)

    # ---- Restoration ----
    ws = wb.create_sheet("Restoration")
    write_df(ws, per_seed_deltas)
    start2 = len(per_seed_deltas) + 3
    ws.cell(row=start2, column=1, value="Seed-level summary (mean +- std, bootstrap 95% CI, N=3 EXPLORATORY)").font = HEADER_FONT
    write_df(ws, seed_level_stats, start_row=start2 + 1)

    # ---- Per_Degradation ----
    ws = wb.create_sheet("Per_Degradation")
    write_df(ws, per_deg_deltas)
    start2 = len(per_deg_deltas) + 3
    ws.cell(row=start2, column=1, value="Per-degradation summary (mean +- std across 3 seeds)").font = HEADER_FONT
    per_deg_summary_flat = per_deg_summary.reset_index()
    per_deg_summary_flat.columns = ["comparison", "degradation", "delta_psnr_mean", "delta_psnr_std",
                                     "delta_ssim_mean", "delta_ssim_std"]
    write_df(ws, per_deg_summary_flat, start_row=start2 + 1)

    # ---- Representation_Probe ----
    ws = wb.create_sheet("Representation_Probe")
    write_df(ws, probe_df)
    start2 = len(probe_df) + 3
    ws.cell(row=start2, column=1, value="Aggregated across seeds (mean +- std)").font = HEADER_FONT
    probe_agg_flat = probe_agg.reset_index()
    probe_agg_flat.columns = ["representation", "accuracy_mean", "accuracy_std", "balanced_accuracy_mean",
                               "balanced_accuracy_std", "macro_f1_mean", "macro_f1_std"]
    write_df(ws, probe_agg_flat, start_row=start2 + 1)
    start3 = start2 + len(probe_agg_flat) + 3
    ws.cell(row=start3, column=1, value="Teacher-student embedding alignment (Model B, C)").font = HEADER_FONT
    align_display = align_df.drop(columns=["per_component_mse"]) if "per_component_mse" in align_df.columns else align_df
    write_df(ws, align_display, start_row=start3 + 1)

    # ---- Conditioning_Statistics ----
    ws = wb.create_sheet("Conditioning_Statistics")
    write_df(ws, cond_df)

    # ---- Bottleneck_Change ----
    ws = wb.create_sheet("Bottleneck_Change")
    write_df(ws, change_df)

    # ---- Random_Control ----
    ws = wb.create_sheet("Random_Control")
    write_df(ws, control_df)
    start2 = len(control_df) + 3
    ws.cell(row=start2, column=1, value="Aggregated by condition").font = HEADER_FONT
    control_agg = control_df.groupby("condition")[["psnr", "ssim"]].agg(["mean", "std"]).reset_index()
    control_agg.columns = ["condition", "psnr_mean", "psnr_std", "ssim_mean", "ssim_std"]
    write_df(ws, control_agg, start_row=start2 + 1)

    # ---- Embedding_Intervention ----
    ws = wb.create_sheet("Embedding_Intervention")
    write_df(ws, interv_df)
    start2 = len(interv_df) + 3
    ws.cell(row=start2, column=1, value="Direction summary (mean across seeds+scenes)").font = HEADER_FONT
    interv_summary = interv_df.groupby(["recipient_degradation", "donor_degradation"])[
        ["psnr_vs_clean", "recipient_normal_psnr", "dist_to_recipient_normal_output",
         "dist_to_donor_normal_output", "relative_output_change"]].mean().reset_index()
    interv_summary["delta_psnr_vs_recipient_normal"] = (
        interv_summary["psnr_vs_clean"] - interv_summary["recipient_normal_psnr"])
    write_df(ws, interv_summary, start_row=start2 + 1)

    # ---- Complexity ----
    ws = wb.create_sheet("Complexity")
    a_row = seed_summary[seed_summary.model == "A"].iloc[0]
    b_row = seed_summary[seed_summary.model == "B"].iloc[0]
    c_row = seed_summary[seed_summary.model == "C"].iloc[0]
    complexity_rows = pd.DataFrame({
        "field": ["params_A", "params_B", "params_C", "extra_params_C_minus_A", "pct_param_overhead_C_vs_A",
                  "extra_params_C_minus_B", "macs_A", "macs_B", "macs_C", "extra_macs_C_minus_A",
                  "pct_macs_overhead_C_vs_A", "conditioning_head_cost", "note"],
        "value": [int(a_row.params), int(b_row.params), int(c_row.params),
                  int(c_row.params) - int(a_row.params),
                  f"{(int(c_row.params) - int(a_row.params)) / a_row.params * 100:.4f}%",
                  int(c_row.params) - int(b_row.params),
                  int(a_row.macs), int(b_row.macs), int(c_row.macs), int(c_row.macs) - int(a_row.macs),
                  f"{(int(c_row.macs) - int(a_row.macs)) / a_row.macs * 100:.4f}%",
                  "2x Linear(16,256) heads (gamma_head, beta_head) + channel-wise affine (elementwise, "
                  "negligible MACs)",
                  "THEORETICAL COMPLEXITY ONLY. NOT an NPU latency claim -- per project finding F1, "
                  "normalization (not MACs) dominates INT8 Hexagon NPU latency."],
    })
    write_df(ws, complexity_rows)

    # ---- Environment ----
    ws = wb.create_sheet("Environment")
    env_rows = pd.DataFrame({
        "field": ["host", "gpu", "conda_env", "cpu_pinning", "thread_limits", "n_parallel_training_processes"],
        "value": ["devon (192.248.10.68)", "RTX 4090", "adair-distill",
                  "taskset -c 0-7,12-31 (CPUs 8-11 excluded: known data-corruption risk)",
                  "OMP_NUM_THREADS=3, MKL_NUM_THREADS=3 (9-way parallel training)", 9],
    })
    write_df(ws, env_rows)

    # ---- GO_NO_GO ----
    ws = wb.create_sheet("GO_NO_GO")
    cb_psnr = seed_level_stats[(seed_level_stats.comparison == "C-B") &
                                (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    ca_psnr = seed_level_stats[(seed_level_stats.comparison == "C-A") &
                                (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    rain_cb = per_deg_summary.loc[("C-B", "rain"), ("delta_psnr", "mean")]
    haze_cb = per_deg_summary.loc[("C-B", "haze"), ("delta_psnr", "mean")]
    noise_cb = per_deg_summary.loc[("C-B", "noise"), ("delta_psnr", "mean")]
    learned_psnr = control_df[control_df.condition == "learned"].psnr.mean()
    random_psnr = control_df[control_df.condition == "random_matched"].psnr.mean()
    go_rows = pd.DataFrame({
        "field": ["decision", "mean_delta_CB_last5_psnr_dB", "CB_all_3_seeds_positive",
                  "mean_delta_CA_last5_psnr_dB", "rain_delta_CB_dB", "haze_delta_CB_dB", "noise_delta_CB_dB",
                  "learned_vs_random_conditioning_gap_dB", "embedding_intervention_causal",
                  "rationale"],
        "value": ["PARTIAL GO", round(cb_psnr["mean"], 3), True, round(ca_psnr["mean"], 3),
                  round(rain_cb, 3), round(haze_cb, 3), round(noise_cb, 3),
                  round(learned_psnr - random_psnr, 3), True,
                  "Model C consistently beats B on the primary metric in all 3 seeds (mean "
                  "+0.27dB, 95% CI excludes zero), and this is causally real: learned "
                  "conditioning beats random/shuffled/zero conditioning by ~3.5dB, and swapping "
                  "in a donor scene's embedding always hurts restoration (-1 to -7.5dB) -- e_S "
                  "demonstrably controls restoration. However the C-B gain is concentrated almost "
                  "entirely in Rain (+0.85dB, recovering much of B's -2.59dB Rain regression), "
                  "while Haze remains essentially unrecovered (-0.06dB) and Noise is flat "
                  "(+0.03dB). C still underperforms baseline A overall (-0.52dB mean, all 3 "
                  "seeds negative). This matches the PARTIAL GO criterion: C~B overall improved "
                  "but strongly recovers specific degradation failures -- warrants "
                  "degradation-specific or multi-depth conditioning as the next step, not a "
                  "wholesale architecture change."],
    })
    write_df(ws, go_rows)

    # ---- embed visualizations ----
    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 480, 300
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 18

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
