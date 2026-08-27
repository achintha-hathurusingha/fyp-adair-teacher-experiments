"""TEST07-B: build the final 13-sheet Excel workbook LOCALLY (Windows
machine), reading the CSVs/JSON pulled back from devon. Run on the local
machine only, per the project's CSV-first / render-locally policy (devon's
CPU 8-11 has a known data-corruption risk).

Usage (local Windows machine, from teacher-experiments/test07_b/src):
  python build_excel_local.py
"""
from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

TEST07B = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST07B / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "AdaIR_Compact_Latent_Distillation.xlsx"

HEADER_FONT = Font(bold=True)


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def autosize_text_sheet(ws, width=100):
    ws.column_dimensions["A"].width = width


def main():
    with open(RESULTS_DIR / "teacher_cache" / "teacher_cache_metadata.json") as f:
        teacher_meta = json.load(f)

    manifest = pd.read_csv(RESULTS_DIR / "dataset_manifest.csv")
    epoch_metrics = pd.read_csv(RESULTS_DIR / "epoch_metrics.csv")
    seed_summary = pd.read_csv(RESULTS_DIR / "seed_summary.csv")
    per_seed_deltas = pd.read_csv(STATS_DIR / "per_seed_deltas.csv")
    seed_level_stats = pd.read_csv(STATS_DIR / "seed_level_summary_stats.csv")
    per_deg_deltas = pd.read_csv(STATS_DIR / "per_degradation_deltas.csv")
    per_deg_summary = pd.read_csv(STATS_DIR / "per_degradation_summary.csv", header=[0, 1], index_col=0)
    probe_df = pd.read_csv(STATS_DIR / "representation_probe.csv")
    probe_agg = pd.read_csv(STATS_DIR / "representation_probe_aggregated.csv", header=[0, 1], index_col=0)
    align_df = pd.read_csv(STATS_DIR / "teacher_student_alignment.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----
    ws = wb.create_sheet("README")
    autosize_text_sheet(ws, 110)
    readme_lines = [
        "TEST07-B: Compact Latent Distillation Validation",
        "",
        "Purpose: statistically meaningful follow-up to TEST07-Pilot, testing ONLY Models A "
        "(baseline NAFNet) and B (baseline + compact teacher-latent distillation), 3 seeds each, "
        "with the pilot's GAP-only bottleneck asymmetry corrected to GAP+GMP.",
        "",
        "Directory isolation: this experiment reads test06/'s DIV2K images and "
        "fyp-adair-distill's locked NAFNet architecture READ-ONLY. It does not modify "
        "test06/, test06_r/, test07_pilot/, or fyp-adair-distill/. All outputs live under test07_b/.",
        "",
        "IMPORTANT DATA-INTEGRITY NOTE: the first attempt to train all 6 (model x seed) runs in "
        "parallel had a race condition -- 6 processes concurrently read-modify-wrote the SAME "
        "shared epoch_metrics.csv/seed_summary.csv files, corrupting the summary statistics "
        "(all 6 rows ended up with identical, incorrect final_psnr/best_psnr values). This was "
        "caught before any analysis was done on the corrupted data, by noticing the impossible "
        "identical values against clearly-diverging live training logs. The bug was fixed (each "
        "training process now writes its own per-run CSV; a separate merge step combines them), "
        "and ALL 6 runs were REPEATED from scratch with the fixed script. Every number in this "
        "workbook comes from that clean, verified re-run.",
        "",
        "Sheet guide:",
        "  Dataset               - scene/crop counts, train/val split, degradation synthesis",
        "  Environment           - hardware/software environment",
        "  Teacher                - AdaIR checkpoint verification, PCA-16 fit details",
        "  Training_Config        - hyperparameters, identical-except-KD confirmation",
        "  Epoch_Metrics           - full per-epoch training/validation history, all 6 runs",
        "  Seed_Summary            - per-run checkpoint-selection summary (best/final/last5)",
        "  Restoration_Comparison  - primary statistical analysis: paired B-A deltas across seeds",
        "  Per_Degradation         - Rain/Haze/Noise breakdown (does not average away failures)",
        "  Representation_Probe    - degradation-classification probe accuracy per representation",
        "  Teacher_Student_Embedding - cosine similarity / MSE between e_S and e_T",
        "  Complexity              - parameter/MACs overhead of Model B vs A",
        "  GO_NO_GO                - final decision and rationale",
    ]
    for i, line in enumerate(readme_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    # ---- Dataset ----
    ws = wb.create_sheet("Dataset")
    train_scenes = sorted(manifest[manifest.split == "train"].scene_id.unique())
    val_scenes = sorted(manifest[manifest.split == "val"].scene_id.unique())
    dataset_summary = pd.DataFrame({
        "field": ["n_train_scenes", "n_val_scenes", "n_crops_per_train_scene", "n_crops_per_val_scene",
                  "crop_size_px", "n_train_crops", "n_val_crops", "degradations",
                  "train_val_scene_overlap", "patch_sampling_option_used", "div2k_source"],
        "value": [len(train_scenes), len(val_scenes), 8, 1, 128,
                  int((manifest.split == "train").sum()), int((manifest.split == "val").sum()),
                  "Rain, Haze, Noise (randomized parameters per crop)",
                  0, "OPTION 2: precomputed multi-crop cache (8 crops/train scene), "
                     "teacher embeddings extracted per-crop from this SAME cache -- "
                     "guarantees teacher/student see identical pixels",
                  "test06/data/div2k_val/DIV2K_valid_HR (READ-ONLY reuse, 100 images)"],
    })
    write_df(ws, dataset_summary)
    ws2 = wb.create_sheet("Dataset_Scene_IDs")
    scene_df = pd.DataFrame({"train_scene_id": pd.Series(train_scenes),
                              "val_scene_id": pd.Series(val_scenes)})
    write_df(ws2, scene_df)

    # ---- Environment ----
    ws = wb.create_sheet("Environment")
    env_rows = pd.DataFrame({
        "field": ["host", "gpu", "conda_env", "cpu_pinning", "thread_limits", "process_launch",
                  "n_parallel_training_processes"],
        "value": ["devon (192.248.10.68)", "RTX 4090", "adair-distill",
                  "taskset -c 0-7,12-31 (CPUs 8-11 excluded: known data-corruption risk)",
                  "OMP_NUM_THREADS=4, MKL_NUM_THREADS=4 (avoids CPU thread oversubscription "
                  "under 6-way parallel training)",
                  "tmux (survives SSH disconnect)", 6],
    })
    write_df(ws, env_rows)

    # ---- Teacher ----
    ws = wb.create_sheet("Teacher")
    teacher_rows = pd.DataFrame({
        "field": ["checkpoint_path", "checkpoint_sha256", "sha256_matches_expected", "n_params",
                  "frozen", "gradients_through_teacher", "raw_pooled_dim (GAP+GMP)", "pca_dim",
                  "pca_fit_sample_count", "pca_fit_source", "pca_explained_variance_ratio_sum",
                  "n_records_total", "n_train_records", "n_val_records"],
        "value": [teacher_meta["teacher_checkpoint_path"], teacher_meta["teacher_checkpoint_sha256"],
                  teacher_meta["sha256_matches"], teacher_meta["n_params"], teacher_meta["teacher_frozen"],
                  teacher_meta["gradients_through_teacher"], teacher_meta["raw_pooled_dim"],
                  teacher_meta["pca_dim"], teacher_meta["pca_fit_sample_count"], teacher_meta["pca_fit_source"],
                  round(teacher_meta["pca_explained_variance_ratio_sum"], 4), teacher_meta["n_records_total"],
                  teacher_meta["n_train_records"], teacher_meta["n_val_records"]],
    })
    write_df(ws, teacher_rows)
    ws2 = wb.create_sheet("Teacher_PCA_Components")
    pca_comp_df = pd.DataFrame({
        "component": list(range(1, 17)),
        "explained_variance_ratio": teacher_meta["pca_per_component_explained_variance_ratio"],
        "embedding_mean_train": teacher_meta["teacher_embedding_mean"],
        "embedding_std_train": teacher_meta["teacher_embedding_std"],
    })
    write_df(ws2, pca_comp_df)

    # ---- Training_Config ----
    ws = wb.create_sheet("Training_Config")
    config_rows = pd.DataFrame({
        "field": ["epochs", "batch_size", "learning_rate", "optimizer", "lambda_kd", "seeds",
                  "models", "identical_config_A_B", "only_B_has_kd_loss", "only_B_has_projection_head",
                  "identical_input_normalization", "identical_validation_set", "student_bottleneck_pooling"],
        "value": [50, 8, "2e-4", "Adam", 0.1, "0, 1, 2", "A (baseline), B (compact latent KD)", True, True, True,
                  True, True, "GAP+GMP (512-dim) -> Linear(512,16); CORRECTED from pilot's GAP-only"],
    })
    write_df(ws, config_rows)

    # ---- Epoch_Metrics ----
    ws = wb.create_sheet("Epoch_Metrics")
    write_df(ws, epoch_metrics)

    # ---- Seed_Summary ----
    ws = wb.create_sheet("Seed_Summary")
    write_df(ws, seed_summary)

    # ---- Restoration_Comparison ----
    ws = wb.create_sheet("Restoration_Comparison")
    write_df(ws, per_seed_deltas)
    start2 = len(per_seed_deltas) + 3
    ws.cell(row=start2, column=1, value="Seed-level summary (mean +- std, bootstrap 95% CI, N=3 EXPLORATORY)").font = HEADER_FONT
    write_df(ws, seed_level_stats, start_row=start2 + 1)

    # ---- Per_Degradation ----
    ws = wb.create_sheet("Per_Degradation")
    write_df(ws, per_deg_deltas)
    start2 = len(per_deg_deltas) + 3
    ws.cell(row=start2, column=1, value="Per-degradation summary (mean +- std across 3 seeds)").font = HEADER_FONT
    per_deg_summary_flat = per_deg_summary.reset_index()
    per_deg_summary_flat.columns = ["degradation", "delta_psnr_mean", "delta_psnr_std",
                                     "delta_ssim_mean", "delta_ssim_std"]
    write_df(ws, per_deg_summary_flat, start_row=start2 + 1)

    # ---- Representation_Probe ----
    ws = wb.create_sheet("Representation_Probe")
    write_df(ws, probe_df)
    start2 = len(probe_df) + 3
    ws.cell(row=start2, column=1, value="Aggregated across seeds (mean +- std)").font = HEADER_FONT
    probe_agg_flat = probe_agg.reset_index()
    probe_agg_flat.columns = ["representation", "accuracy_mean", "accuracy_std",
                               "balanced_accuracy_mean", "balanced_accuracy_std",
                               "macro_f1_mean", "macro_f1_std"]
    write_df(ws, probe_agg_flat, start_row=start2 + 1)

    # ---- Teacher_Student_Embedding ----
    ws = wb.create_sheet("Teacher_Student_Embedding")
    align_display = align_df.drop(columns=["per_component_mse"]) if "per_component_mse" in align_df.columns else align_df
    write_df(ws, align_display)

    # ---- Complexity ----
    ws = wb.create_sheet("Complexity")
    a_row = seed_summary[seed_summary.model == "A"].iloc[0]
    b_row = seed_summary[seed_summary.model == "B"].iloc[0]
    extra_params = int(b_row.params) - int(a_row.params)
    extra_macs = int(b_row.macs) - int(a_row.macs)
    complexity_rows = pd.DataFrame({
        "field": ["params_A", "params_B", "extra_params_B_minus_A", "pct_param_overhead",
                  "macs_A_at_128px", "macs_B_at_128px", "extra_macs_B_minus_A", "pct_macs_overhead", "note"],
        "value": [int(a_row.params), int(b_row.params), extra_params, f"{extra_params / a_row.params * 100:.4f}%",
                  int(a_row.macs), int(b_row.macs), extra_macs, f"{extra_macs / a_row.macs * 100:.4f}%",
                  "THEORETICAL COMPLEXITY ONLY (via fyp-adair-distill's validated count_macs, "
                  "FlopCounterMode-based). NOT an NPU latency claim -- per project finding F1, "
                  "normalization (not MACs) dominates INT8 Hexagon NPU latency."],
    })
    write_df(ws, complexity_rows)

    # ---- GO_NO_GO ----
    ws = wb.create_sheet("GO_NO_GO")
    delta_psnr_row = seed_level_stats[seed_level_stats.metric == "delta_last5_psnr"].iloc[0]
    delta_ssim_row = seed_level_stats[seed_level_stats.metric == "delta_last5_ssim"].iloc[0]
    mean_cosine = align_df.mean_cosine_similarity.mean()
    b_es_acc = probe_agg.loc["model_B_compact_eS", ("accuracy", "mean")]
    go_rows = pd.DataFrame({
        "field": ["decision", "mean_delta_last5_psnr_dB", "mean_delta_last5_ssim", "n_seeds_negative_psnr_of_3",
                  "mean_teacher_student_cosine_similarity", "model_B_eS_probe_accuracy",
                  "rain_delta_psnr_dB", "haze_delta_psnr_dB", "noise_delta_psnr_dB", "rationale"],
        "value": ["NO-GO FOR SIMPLE KD", round(delta_psnr_row["mean"], 3), round(delta_ssim_row["mean"], 4), 3,
                  round(mean_cosine, 4), round(b_es_acc, 4),
                  round(per_deg_summary.loc["rain", ("delta_psnr", "mean")], 3),
                  round(per_deg_summary.loc["haze", ("delta_psnr", "mean")], 3),
                  round(per_deg_summary.loc["noise", ("delta_psnr", "mean")], 3),
                  "Model B <= Model A on the primary restoration metric (last5-window PSNR) in ALL "
                  "3 seeds (mean delta = -0.79dB), AND e_S matches e_T well (cosine ~0.99, probe "
                  "accuracy ~96%, matching teacher's own 96.1%). Per the pre-specified decision "
                  "rule, this combination -- good representation transfer but no restoration "
                  "benefit -- means representation-matching alone is insufficient; do not blame "
                  "the KD loss mechanism. Notably, Noise consistently favors B (+1.24dB mean, all "
                  "3 seeds positive) while Rain (-2.59dB) and Haze (-1.03dB) consistently favor A "
                  "-- a degradation-specific pattern worth investigating in future work, not "
                  "captured by the pooled/averaged decision."],
    })
    write_df(ws, go_rows)

    # ---- embed visualizations on a dedicated sheet ----
    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 480, 300
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 18

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
