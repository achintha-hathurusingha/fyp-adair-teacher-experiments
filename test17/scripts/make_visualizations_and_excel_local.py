"""TEST17: build visualizations and the final Excel workbook LOCALLY, from
the real trained-model results pulled back from devon (CSV-first /
render-locally policy)."""
from __future__ import annotations

import json
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST17 = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST17 / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "Snapdragon_NAFNet_F2_Deployment.xlsx"
HEADER_FONT = Font(bold=True)
MODEL_NAMES = ["A", "N", "F2", "NF2"]
COLORS = {"A": "#1f77b4", "N": "#2ca02c", "F2": "#ff7f0e", "NF2": "#d62728"}


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def main():
    seed_summary = pd.read_csv(RESULTS_DIR / "seed_summary.csv")
    epoch_metrics = pd.read_csv(RESULTS_DIR / "epoch_metrics.csv")
    full_graph = pd.read_csv(STATS_DIR / "full_graph_benchmark.csv")
    int8 = pd.read_csv(STATS_DIR / "int8_benchmark.csv")
    hotspots_top10 = pd.read_csv(STATS_DIR / "layer_hotspots_top10.csv")
    hotspots_bucket = pd.read_csv(STATS_DIR / "layer_hotspots_by_bucket.csv")
    rep_probe = pd.read_csv(STATS_DIR / "representation_probe.csv")
    teacher_align = pd.read_csv(STATS_DIR / "teacher_alignment.csv")
    with open(RESULTS_DIR / "export_manifest.json") as f:
        export_manifest = {e["name"]: e for e in json.load(f)}

    VIZ_DIR.mkdir(parents=True, exist_ok=True)

    quality = seed_summary.groupby("model")[[
        "last5_mean_psnr", "last5_mean_ssim", "last5_mean_rain_psnr", "last5_mean_haze_psnr",
        "last5_mean_noise_psnr", "last5_mean_rain_ssim", "last5_mean_haze_ssim", "last5_mean_noise_ssim",
    ]].agg(["mean", "std"])
    quality.columns = ["_".join(c) for c in quality.columns]
    quality = quality.reset_index().set_index("model").loc[MODEL_NAMES].reset_index()
    quality.to_csv(STATS_DIR / "quality_summary.csv", index=False)

    master_rows = []
    for name in MODEL_NAMES:
        q = quality[quality.model == name].iloc[0]
        g = full_graph[full_graph.name == name].iloc[0]
        e = export_manifest[name]
        i8 = int8[int8.name == name].iloc[0]
        master_rows.append({
            "Model": name, "PSNR_dB": q.last5_mean_psnr_mean, "PSNR_std": q.last5_mean_psnr_std,
            "SSIM": q.last5_mean_ssim_mean, "SSIM_std": q.last5_mean_ssim_std,
            "Params": int(seed_summary[seed_summary.model == name].params.iloc[0]),
            "MACs": int(seed_summary[seed_summary.model == name].macs.iloc[0]),
            "Model_MB": round(e["model_size_bytes"] / 1e6, 2),
            "NPU_latency_FP32_ms": g.latency_mean_ms, "NPU_latency_INT8_ms": i8.latency_ms,
            "Peak_memory_MB": round(g.peak_memory_bytes / 1e6, 2),
            "INT8_PSNR_dB": i8.int8_psnr, "INT8_SSIM": i8.int8_ssim,
            "CPU_fallback": bool(g.any_cpu_fallback), "NPU_only": bool(g.npu_only),
        })
    master_df = pd.DataFrame(master_rows)
    master_df.to_csv(STATS_DIR / "master_quality_hardware_table.csv", index=False)

    # ---- Viz 1: quality with error bars ----
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(master_df.Model, master_df.PSNR_dB, yerr=master_df.PSNR_std,
           color=[COLORS[n] for n in master_df.Model], capsize=5)
    ax.set_ylabel("Last-5-epoch mean validation PSNR (dB), 3-seed mean +- std")
    ax.set_title("TEST17: Restoration Quality (FP32, real trained weights)")
    ax.set_ylim(15, 30)
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "01_quality_psnr.png", dpi=150)
    plt.close(fig)

    # ---- Viz 2: per-degradation PSNR grouped ----
    fig, ax = plt.subplots(figsize=(10, 5))
    degs = ["rain", "haze", "noise"]
    x = range(len(MODEL_NAMES))
    width = 0.25
    for i, deg in enumerate(degs):
        vals = [quality[quality.model == n][f"last5_mean_{deg}_psnr_mean"].iloc[0] for n in MODEL_NAMES]
        ax.bar([xi + (i - 1) * width for xi in x], vals, width, label=deg.capitalize())
    ax.set_xticks(list(x))
    ax.set_xticklabels(MODEL_NAMES)
    ax.set_ylabel("PSNR (dB)")
    ax.set_title("Per-Degradation PSNR by Model")
    ax.legend()
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "02_per_degradation_psnr.png", dpi=150)
    plt.close(fig)

    # ---- Viz 3: FP32 latency ----
    fig, ax = plt.subplots(figsize=(8, 5))
    bars = ax.bar(full_graph.name, full_graph.latency_mean_ms, color=[COLORS[n] for n in full_graph.name],
                   yerr=full_graph.latency_std_ms, capsize=4)
    ax.set_yscale("log")
    ax.set_ylabel("Mean NPU latency (ms), log scale, n=100")
    ax.set_title("TEST17: Full-Graph FP32 NPU Latency (real trained weights)")
    for bar, v in zip(bars, full_graph.latency_mean_ms):
        ax.text(bar.get_x() + bar.get_width() / 2, v * 1.15, f"{v:.0f}ms", ha="center")
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "03_fp32_latency.png", dpi=150)
    plt.close(fig)

    # ---- Viz 4: FP32 vs INT8 latency ----
    fig, ax = plt.subplots(figsize=(9, 5))
    x = range(len(MODEL_NAMES))
    fp32_vals = [full_graph[full_graph.name == n].iloc[0].latency_mean_ms for n in MODEL_NAMES]
    int8_vals = [int8[int8.name == n].iloc[0].latency_ms for n in MODEL_NAMES]
    ax.bar([i - 0.175 for i in x], fp32_vals, 0.35, label="FP32", color="#1f77b4")
    ax.bar([i + 0.175 for i in x], int8_vals, 0.35, label="INT8", color="#ff7f0e")
    ax.set_yscale("log")
    ax.set_xticks(list(x))
    ax.set_xticklabels(MODEL_NAMES)
    ax.set_ylabel("Latency (ms), log scale")
    ax.set_title("FP32 vs INT8 Latency")
    ax.legend()
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "04_fp32_vs_int8.png", dpi=150)
    plt.close(fig)

    # ---- Viz 5: Pareto latency vs PSNR ----
    fig, ax = plt.subplots(figsize=(8, 6))
    for _, row in master_df.iterrows():
        ax.scatter(row.NPU_latency_FP32_ms, row.PSNR_dB, s=150, color=COLORS[row.Model])
        ax.annotate(row.Model, (row.NPU_latency_FP32_ms, row.PSNR_dB), textcoords="offset points", xytext=(8, 8))
    ax.set_xscale("log")
    ax.set_xlabel("Full-graph NPU latency (ms), log scale")
    ax.set_ylabel("Validation PSNR (dB)")
    ax.set_title("Quality/Latency Pareto (FP32)")
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "05_pareto_latency_psnr.png", dpi=150)
    plt.close(fig)

    # ---- Viz 6: Pareto latency vs SSIM ----
    fig, ax = plt.subplots(figsize=(8, 6))
    for _, row in master_df.iterrows():
        ax.scatter(row.NPU_latency_FP32_ms, row.SSIM, s=150, color=COLORS[row.Model])
        ax.annotate(row.Model, (row.NPU_latency_FP32_ms, row.SSIM), textcoords="offset points", xytext=(8, 8))
    ax.set_xscale("log")
    ax.set_xlabel("Full-graph NPU latency (ms), log scale")
    ax.set_ylabel("Validation SSIM")
    ax.set_title("Quality/Latency Pareto (FP32, SSIM)")
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "06_pareto_latency_ssim.png", dpi=150)
    plt.close(fig)

    # ---- Viz 7: latency vs memory ----
    fig, ax = plt.subplots(figsize=(8, 6))
    for _, row in master_df.iterrows():
        ax.scatter(row.NPU_latency_FP32_ms, row.Peak_memory_MB, s=150, color=COLORS[row.Model])
        ax.annotate(row.Model, (row.NPU_latency_FP32_ms, row.Peak_memory_MB), textcoords="offset points", xytext=(8, 8))
    ax.set_xscale("log")
    ax.set_xlabel("Full-graph NPU latency (ms), log scale")
    ax.set_ylabel("Peak inference memory (MB)")
    ax.set_title("Latency vs Peak Memory")
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(VIZ_DIR / "07_latency_vs_memory.png", dpi=150)
    plt.close(fig)

    print(f"wrote 7 visualizations to {VIZ_DIR}")

    # ================= Excel workbook =================
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 130
    readme = [
        "TEST17: Hardware-Aware Student Validation (Normalization Surgery + F2 Conditioning, Trained)",
        "",
        "Core question: can the validated F2 degradation-conditioned operator be combined with the "
        "NPU-friendly normalization-surgery backbone (this time trained for real) and retain useful "
        "restoration quality while keeping most of the NPU latency win?",
        "",
        "HEADLINE FINDINGS:",
        "1. Quality: A=27.31dB, F2=27.04dB (both stable, seed std<0.21). N=23.45dB, seed-UNSTABLE "
        "(std=1.78) -- a genuine training-divergence event (loss spike epoch 27) was found in one "
        "seed's own epoch curve. N+F2=25.04dB, seed-STABLE (std=0.12) -- F2's conditioning mechanism "
        "stabilizes training on the fast backbone, not just adds quality.",
        "2. Per-degradation: F2 vs A -- Rain-1.13, Haze-0.89, Noise+1.20. N vs A -- Rain-4.31, "
        "Haze-5.20 (worst), Noise-2.09. N+F2 vs A -- Rain-3.25, Haze-2.02, Noise-1.55. N+F2 recovers "
        "most of N's Haze loss specifically (+3.17dB over N alone).",
        "3. Representation: F2 and N+F2's e_D are nearly identical quality (probe 96.6%/cosine 0.989 "
        "vs 95.6%/0.973) -- N+F2's quality gap vs F2 is NOT explained by worse degradation "
        "representation learning; it is a backbone effect.",
        "4. Hardware (FP32, real trained weights): A=10,560ms, N=3,777ms, F2=10,610ms, N+F2=3,844ms "
        "(+66ms/+1.7% over N -- confirms N+F2 approx= N, not approx= F2). CORRECTS TEST16: with real "
        "trained weights the normalization-surgery speedup is ~2.80x (A/N), not TEST16's untrained "
        "~24x, which was an artifact of identity-valued (weight=1,bias=0) untrained affine-norm "
        "layers being constant-folded away by the NPU compiler.",
        "5. INT8: A=2,076ms/27.87dB, N=1,635ms/25.01dB, F2=2,557ms/26.22dB, N+F2=1,678ms/25.94dB "
        "(12-sample subset each -- some INT8 PSNR readings are HIGHER than FP32, almost certainly a "
        "small-sample artifact from a different random draw of images, not genuine quantization "
        "improvement). All 4 remain 100% NPU, zero fallback, under INT8 too.",
        "6. All 4 models compiled and ran 100% NPU with ZERO CPU/GPU fallback in both FP32 and INT8.",
        "",
        "Device: Snapdragon 8 Elite QRD (SM8750, Hexagon V79) -- identical to TEST15/16. "
        "Runtime: QNN context binary. Resolution: 256x256 production export path, opset 17.",
        "",
        "Caveats: INT8 quality figures use a single seed's (seed=0) best checkpoint and a distinct "
        "12-image sample (different seed than the FP32 3-seed validation) -- a directional signal, "
        "not a precise benchmark-grade comparison. N's instability (one seed's training divergence) "
        "means its 3-seed mean PSNR should be read with that variance in mind, not as a single clean "
        "number.",
    ]
    for i, line in enumerate(readme, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    ws = wb.create_sheet("Environment")
    write_df(ws, pd.DataFrame({
        "field": ["target_device", "chipset", "npu", "runtime", "onnx_opset", "input_resolution",
                  "epochs", "optimizer", "lr", "batch_size", "seeds"],
        "value": ["Snapdragon 8 Elite QRD", "SM8750", "Hexagon V79", "qnn_context_binary", 17,
                  "256x256 (export) / 128x128 (training crops)", 50, "Adam", "2e-4", 8, "0,1,2"],
    }))

    ws = wb.create_sheet("Models")
    write_df(ws, pd.DataFrame({
        "Model": ["A", "N", "F2", "NF2"],
        "Description": [
            "Original baseline NAFNet, layernorm2d backbone, no conditioning.",
            "Normalization surgery: layernorm2d -> affine_clamp, no conditioning.",
            "Original layernorm2d backbone + F2's validated rank-2 low-rank conditional operator.",
            "F2's exact mechanism, unchanged, applied to the affine_clamp (norm-surgery) backbone.",
        ],
        "Trained": [True, True, True, True],
    }))

    ws = wb.create_sheet("Dataset")
    write_df(ws, pd.DataFrame({
        "field": ["source", "split", "train_scenes", "val_scenes", "degradations", "crop_size"],
        "value": ["TEST12/TEST07-B dataset (DIV2K-derived, reused read-only)", "80/20 scene-disjoint",
                   80, 20, "Rain, Haze, Noise", "128x128"],
    }))

    ws = wb.create_sheet("Training_Config")
    write_df(ws, pd.DataFrame({
        "field": ["epochs", "optimizer", "lr", "batch_size", "loss_A_N", "loss_F2_NF2", "lambda_kd"],
        "value": [50, "Adam", "2e-4", 8, "L1(output, clean)",
                   "L1(output, clean) + 0.1 * MSE(e_S, e_T)", 0.1],
    }))

    ws = wb.create_sheet("Epoch_Metrics")
    write_df(ws, epoch_metrics)

    ws = wb.create_sheet("Seed_Summary")
    write_df(ws, seed_summary)

    ws = wb.create_sheet("Restoration")
    write_df(ws, master_df[["Model", "PSNR_dB", "PSNR_std", "SSIM", "SSIM_std"]])

    ws = wb.create_sheet("Per_Degradation")
    per_deg_cols = ["model"] + [c for c in quality.columns if "rain" in c or "haze" in c or "noise" in c]
    write_df(ws, quality[per_deg_cols])

    ws = wb.create_sheet("Representation")
    write_df(ws, rep_probe)
    r0 = len(rep_probe) + 3
    write_df(ws, teacher_align, start_row=r0)

    ws = wb.create_sheet("PyTorch_Validation")
    write_df(ws, seed_summary[["model", "seed", "final_psnr", "best_psnr", "last5_mean_psnr",
                                "last5_mean_ssim", "any_nan_or_inf", "params", "macs"]])

    ws = wb.create_sheet("ONNX_Export")
    export_rows = pd.DataFrame([{
        "name": n, "export_status": export_manifest[n]["export_status"],
        "node_count": export_manifest[n]["node_count"],
        "model_size_MB": round(export_manifest[n]["model_size_bytes"] / 1e6, 2),
        "n_conv": export_manifest[n]["n_conv"], "n_conv_dynamic_weight": export_manifest[n]["n_conv_dynamic_weight"],
        "all_conv_weights_constant": export_manifest[n]["all_conv_weights_constant"],
    } for n in MODEL_NAMES])
    write_df(ws, export_rows)

    ws = wb.create_sheet("Compilation")
    write_df(ws, full_graph[["name", "compile_success", "profile_success", "device", "compute_units_used",
                              "npu_only", "any_cpu_fallback", "any_gpu_fallback", "n_layers"]])

    ws = wb.create_sheet("Latency")
    write_df(ws, master_df[["Model", "NPU_latency_FP32_ms", "NPU_latency_INT8_ms"]])
    r0 = 8
    write_df(ws, full_graph, start_row=r0)

    ws = wb.create_sheet("Memory")
    write_df(ws, master_df[["Model", "Peak_memory_MB", "Model_MB"]])

    ws = wb.create_sheet("Layer_Hotspots")
    write_df(ws, hotspots_top10)
    r0 = len(hotspots_top10) + 3
    write_df(ws, hotspots_bucket, start_row=r0)

    ws = wb.create_sheet("INT8")
    write_df(ws, int8)

    ws = wb.create_sheet("Pareto")
    write_df(ws, master_df)

    ws = wb.create_sheet("GO_NO_GO")
    write_df(ws, pd.DataFrame({
        "question": [
            "Does F2 survive normalization surgery (does N+F2 stay close to N latency)?",
            "Does N+F2 retain useful restoration quality vs A?",
            "Is normalization surgery alone (N) safe to deploy without F2?",
            "Overall decision",
        ],
        "answer": [
            "YES -- N+F2=3,844ms vs N=3,777ms, only +1.7%. The operator adds negligible latency "
            "regardless of backbone (also true on the slow backbone: F2 vs A is +0.5%).",
            "PARTIALLY -- N+F2 trades ~2.3dB vs A (25.04 vs 27.31dB) for a ~2.80x latency win. "
            "Recovers most of N's worst degradation loss (Haze) but does not fully close the gap.",
            "NO -- N alone is quality-unstable (std=1.78 across seeds, one seed shows a genuine "
            "training-divergence event). F2's conditioning appears necessary for stable training on "
            "this backbone, not just for extra quality.",
            "PARTIAL GO for N+F2 -- real, substantial ~2.80x latency win (corrected from TEST16's "
            "untrained ~24x estimate), non-trivial but bounded ~2.3dB quality cost, and critically "
            "more stable than the normalization-surgery backbone alone. Recommended as the primary "
            "deployment candidate pending a wider/deeper backbone sweep to see if the quality gap "
            "can be narrowed without losing the latency advantage.",
        ],
    }))

    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 560, 380
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 22

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
