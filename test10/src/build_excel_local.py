"""TEST10: build the final 14-sheet Excel workbook LOCALLY (Windows
machine), reading CSVs pulled back from devon. Run locally only, per the
project's CSV-first / render-locally policy.

Usage (local Windows machine, from teacher-experiments/test10/src):
  python build_excel_local.py
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST10 = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST10 / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "AdaIR_Restoration_Trajectory_Distillation.xlsx"

HEADER_FONT = Font(bold=True)


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def autosize_text_sheet(ws, width=115):
    ws.column_dimensions["A"].width = width


def main():
    epoch_metrics = pd.read_csv(RESULTS_DIR / "epoch_metrics.csv")
    seed_summary = pd.read_csv(RESULTS_DIR / "seed_summary.csv")
    per_seed_deltas = pd.read_csv(STATS_DIR / "per_seed_deltas.csv")
    seed_level_stats = pd.read_csv(STATS_DIR / "seed_level_summary_stats.csv")
    per_deg_deltas = pd.read_csv(STATS_DIR / "per_degradation_deltas.csv")
    per_deg_summary = pd.read_csv(STATS_DIR / "per_degradation_summary.csv", header=[0, 1], index_col=[0, 1])
    stage_alignment = pd.read_csv(STATS_DIR / "stage_alignment.csv")
    collapse_diag = pd.read_csv(STATS_DIR / "collapse_diagnostics.csv")
    final_align = pd.read_csv(STATS_DIR / "final_embedding_alignment.csv")
    residual_df = pd.read_csv(STATS_DIR / "residual_analysis_summary.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----
    ws = wb.create_sheet("README")
    autosize_text_sheet(ws)
    readme_lines = [
        "TEST10: Restoration Trajectory Distillation",
        "",
        "Research question: does the student need to learn the teacher's restoration "
        "TRANSFORMATION TRAJECTORY (not just its final compact latent state) to convert "
        "degradation knowledge into restoration quality, especially for Haze -- which "
        "TEST08-C's bottleneck FiLM and TEST09's deeper-FiLM/low-rank mechanisms could "
        "never reliably fix?",
        "",
        "CRITICAL FINDING: the trajectory-distillation mechanism, as specified (freely "
        "jointly-trained per-stage projection heads on both teacher and student sides, "
        "MSE loss on L2-normalized embeddings, no negative pairs or variance regularizer), "
        "suffered TOTAL REPRESENTATIONAL COLLAPSE. All 3 trajectory stages, all 3 seeds, "
        "show inter-sample pairwise cosine similarity > 0.999 -- the projected embeddings "
        "barely vary across different degradation/scene inputs at all. The near-zero "
        "training loss observed from epoch ~10 onward reflects this collapse, NOT genuine "
        "trajectory alignment. This directly invalidates any restoration-quality "
        "conclusion attributable to 'the student learned the teacher's trajectory' -- see "
        "the GO_NO_GO sheet and the full report for the methodology fix needed before this "
        "hypothesis can be properly tested.",
        "",
        "For contrast, the ESTABLISHED final 16-dim KD embedding (same MSE-after-KD "
        "mechanism validated since TEST07-B, using a FIXED leakage-safe PCA teacher "
        "target rather than a freely jointly-trained one) did NOT collapse in this same "
        "run (cosine ~0.78-0.80, consistent with prior experiments) -- isolating the "
        "collapse specifically to the new trajectory heads' training discipline, not a "
        "general problem with this run or codebase.",
        "",
        "Directory isolation: reads test07_b's dataset/teacher-cache and "
        "teacher-experiments/scripts/instrument.py READ-ONLY; imports fyp-adair-distill's "
        "locked NAFNet and the vendored AdaIR checkpoint READ-ONLY. Does not modify any "
        "prior experiment directory. All outputs live under test10/.",
        "",
        "Models: A (baseline), F (TEST09's best mechanism reproduced, low-rank rank=2 "
        "per this task's spec), G (F + trajectory distillation at 3 stages matched to "
        "teacher AFLB1/2/3 by SPATIAL RESOLUTION, per the Phase-0 audit -- see "
        "Teacher_Stages sheet).",
        "",
        "Sheet guide:",
        "  Teacher_Stages    - Phase 0 audit: available stages, chosen correspondence",
        "  Dataset            - reused TEST07-B/08-C/09 dataset (reference only)",
        "  Training_Config     - hyperparameters, loss formulation per model",
        "  Epoch_Metrics        - full per-epoch history, all 9 runs, incl. per-stage traj loss",
        "  Seed_Summary          - per-run checkpoint-selection summary",
        "  Restoration            - primary statistical analysis: G-F, G-A, F-A",
        "  Per_Degradation         - Rain/Haze/Noise breakdown for all comparisons",
        "  Stage_Alignment          - cosine/normalized-MSE per stage PLUS collapse diagnostics",
        "  Trajectory_Loss           - per-stage training loss trend (see Epoch_Metrics for raw data)",
        "  Residual_Analysis          - Phase 11-12: output-change and teacher/student residual analysis",
        "  Complexity                  - parameter/MAC overhead, deployable graph only",
        "  GO_NO_GO                     - final decision, dominated by the collapse finding",
    ]
    for i, line in enumerate(readme_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    # ---- Teacher_Stages ----
    ws = wb.create_sheet("Teacher_Stages")
    stages_df = pd.DataFrame({
        "trajectory_stage": [0, 1, 2],
        "teacher_tensor": ["AFLB1.aflb_out", "AFLB2.aflb_out", "AFLB3.aflb_out"],
        "teacher_shape_128px": ["(1,384,16,16)", "(1,192,32,32)", "(1,96,64,64)"],
        "student_tensor": ["net.decoders[0] output", "net.decoders[1] output", "net.decoders[2] output"],
        "student_shape_128px": ["(1,128,16,16)", "(1,64,32,32)", "(1,32,64,64)"],
        "matched_by": ["spatial resolution (16x16)"] * 1 + ["spatial resolution (32x32)"] * 1 +
                      ["spatial resolution (64x64)"],
        "projection": ["GAP+GMP -> Linear -> 32-dim (both sides)"] * 3,
    })
    write_df(ws, stages_df)

    # ---- Dataset ----
    ws = wb.create_sheet("Dataset")
    dataset_rows = pd.DataFrame({
        "field": ["source", "n_train_scenes", "n_val_scenes", "crops_per_train_scene", "crop_size_px",
                  "degradations"],
        "value": ["REUSED from test07_b/ (READ-ONLY, identical split/crops to TEST08-C/09)",
                  80, 20, 8, 128, "Rain, Haze, Noise"],
    })
    write_df(ws, dataset_rows)

    # ---- Training_Config ----
    ws = wb.create_sheet("Training_Config")
    config_rows = pd.DataFrame({
        "field": ["epochs", "batch_size", "learning_rate", "optimizer", "lambda_kd", "lambda_traj",
                  "stage_weights", "lowrank_rank", "traj_dim", "seeds",
                  "loss_A", "loss_F", "loss_G", "delta_transformation_ablation_(Phase 4)"],
        "value": [50, 8, "2e-4", "Adam", 0.1, 0.1, "1/3 per stage (equal)", 2, 32, "0, 1, 2",
                  "L_restore", "L_restore + 0.1*L_KD",
                  "L_restore + 0.1*L_KD + 0.1*L_traj (L_traj = mean of 3 stage MSE-after-L2norm losses)",
                  "NOT RUN -- explicitly optional per spec ('only run after basic trajectory model is "
                  "working'); basic model was found to have collapsed, so this ablation was skipped "
                  "rather than compounding an already-invalid mechanism"],
    })
    write_df(ws, config_rows)

    # ---- Epoch_Metrics ----
    ws = wb.create_sheet("Epoch_Metrics")
    write_df(ws, epoch_metrics)

    # ---- Seed_Summary ----
    ws = wb.create_sheet("Seed_Summary")
    write_df(ws, seed_summary)

    # ---- Restoration ----
    ws = wb.create_sheet("Restoration")
    write_df(ws, per_seed_deltas)
    start2 = len(per_seed_deltas) + 3
    ws.cell(row=start2, column=1, value="Seed-level summary (mean +- std, bootstrap 95% CI, N=3 EXPLORATORY)").font = HEADER_FONT
    write_df(ws, seed_level_stats, start_row=start2 + 1)

    # ---- Per_Degradation ----
    ws = wb.create_sheet("Per_Degradation")
    write_df(ws, per_deg_deltas)
    start2 = len(per_deg_deltas) + 3
    ws.cell(row=start2, column=1, value="Per-degradation summary (mean +- std across 3 seeds)").font = HEADER_FONT
    per_deg_summary_flat = per_deg_summary.reset_index()
    per_deg_summary_flat.columns = ["comparison", "degradation", "delta_psnr_mean", "delta_psnr_std",
                                     "delta_ssim_mean", "delta_ssim_std"]
    write_df(ws, per_deg_summary_flat, start_row=start2 + 1)

    # ---- Stage_Alignment ----
    ws = wb.create_sheet("Stage_Alignment")
    write_df(ws, stage_alignment)
    start2 = len(stage_alignment) + 3
    ws.cell(row=start2, column=1, value="COLLAPSE DIAGNOSTICS (critical -- see README)").font = HEADER_FONT
    write_df(ws, collapse_diag, start_row=start2 + 1)
    start3 = start2 + len(collapse_diag) + 3
    ws.cell(row=start3, column=1, value="Final 16-dim KD embedding alignment (NOT collapsed, for contrast)").font = HEADER_FONT
    write_df(ws, final_align, start_row=start3 + 1)

    # ---- Trajectory_Loss ----
    ws = wb.create_sheet("Trajectory_Loss")
    traj_cols = ["model", "seed", "epoch", "train_traj_loss", "train_stage0_loss", "train_stage1_loss",
                 "train_stage2_loss"]
    g_traj = epoch_metrics[epoch_metrics.model == "G"][traj_cols]
    write_df(ws, g_traj)

    # ---- Residual_Analysis ----
    ws = wb.create_sheet("Residual_Analysis")
    write_df(ws, residual_df)

    # ---- Complexity ----
    ws = wb.create_sheet("Complexity")
    a_row = seed_summary[seed_summary.model == "A"].iloc[0]
    f_row = seed_summary[seed_summary.model == "F"].iloc[0]
    g_row = seed_summary[seed_summary.model == "G"].iloc[0]
    complexity_rows = pd.DataFrame({
        "field": ["params_A", "params_F", "params_G_deployable", "params_G_training_incl_traj_heads",
                  "extra_params_G_deployable_vs_F", "traj_heads_params_discarded_at_inference",
                  "teacher_params_never_saved_to_student", "macs_A", "macs_F", "macs_G", "note"],
        "value": [int(a_row.params), int(f_row.params), int(g_row.params) - 14432, int(g_row.params),
                  0, 14432, 28784824, int(a_row.macs), int(f_row.macs), int(g_row.macs),
                  "MACs for F and G are IDENTICAL -- confirms traj_heads contribute ZERO ops to the "
                  "actual forward() computation graph (they are only invoked by forward_trajectory(), "
                  "training-only). See verify_inference_graph.py: bit-identical output confirmed. "
                  "THEORETICAL COMPLEXITY ONLY, not an NPU latency claim."],
    })
    write_df(ws, complexity_rows)

    # ---- Environment ----
    ws = wb.create_sheet("Environment")
    env_rows = pd.DataFrame({
        "field": ["host", "gpu", "conda_env", "cpu_pinning", "wave1_A_F", "wave2_G",
                  "G_gpu_memory_per_process", "G_epoch_time_note"],
        "value": ["devon (192.248.10.68)", "RTX 4090", "adair-distill",
                  "taskset -c 0-7,12-31", "6-way concurrent (~2GB/process)",
                  "3-way concurrent (~4GB/process, loads frozen 28.8M-param AdaIR teacher online)",
                  "~3988 MiB", "~36s/epoch standalone, ~84s/epoch under 3-way GPU contention "
                  "(no meaningful slowdown from the teacher itself -- confirmed via standalone timing "
                  "test -- slowdown is purely concurrency contention)"],
    })
    write_df(ws, env_rows)

    # ---- GO_NO_GO ----
    ws = wb.create_sheet("GO_NO_GO")
    gf_psnr = seed_level_stats[(seed_level_stats.comparison == "G-F") &
                                (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    haze_gf = per_deg_summary.loc[("G-F", "haze"), ("delta_psnr", "mean")]
    go_rows = pd.DataFrame({
        "field": ["decision", "mean_delta_GF_psnr_dB", "GF_all_3_seeds_same_sign", "haze_delta_GF_dB",
                  "trajectory_representation_collapsed", "inference_graph_verified_clean", "rationale"],
        "value": ["NO-GO (METHODOLOGY FAILURE, not evidence against H_TRAJ)", round(gf_psnr["mean"], 3),
                  "Yes, all 3 seeds negative", round(haze_gf, 3), True, True,
                  "Model G underperforms Model F in all 3 seeds (mean -0.334dB) and does not improve "
                  "Haze (-0.176dB). However this CANNOT be interpreted as 'trajectory distillation "
                  "doesn't work' -- the trajectory representation itself collapsed to a near-constant "
                  "vector (pairwise cosine >0.999 across all stages/seeds), meaning the student never "
                  "received genuine per-sample trajectory supervision. The observed negative delta is "
                  "more likely explained by the collapsed loss acting as unhelpful gradient noise "
                  "during training, not by 'having learned the trajectory hurt restoration.' The "
                  "correct verdict is: this specific implementation (freely jointly-trained projection "
                  "heads, no collapse-prevention mechanism) failed methodologically and must be fixed "
                  "(e.g., fixed/frozen or leakage-safe-PCA-fit stage projections, matching the "
                  "successful discipline already used for the final 16-dim KD embedding, which did "
                  "NOT collapse in this same run) before H_TRAJ can be properly tested. Separately, "
                  "residual analysis found the teacher's OWN Haze residual-to-GT is worse than "
                  "baseline Model A's, suggesting part of the persistent Haze failure across "
                  "TEST08-C/09/10 may reflect a genuine teacher-quality ceiling for this synthetic "
                  "Haze recipe, not purely a student-mechanism limitation."],
    })
    write_df(ws, go_rows)

    # ---- Visualizations ----
    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 480, 300
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 18

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
