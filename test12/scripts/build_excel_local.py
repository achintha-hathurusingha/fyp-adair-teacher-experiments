"""TEST12: build the final 15-sheet Excel workbook LOCALLY. Run on the
local machine only, per the project's CSV-first / render-locally policy.

Usage (local Windows machine, from teacher-experiments/test12/scripts):
  python build_excel_local.py
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

TEST12 = Path(__file__).resolve().parent.parent
RESULTS_DIR = TEST12 / "results"
STATS_DIR = RESULTS_DIR / "statistics"
VIZ_DIR = RESULTS_DIR / "visualizations"
OUT_XLSX = RESULTS_DIR / "AdaIR_Feature_Conditioned_Operator.xlsx"

HEADER_FONT = Font(bold=True)


def write_df(ws, df, start_row=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                cell.font = HEADER_FONT
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 45)


def autosize_text_sheet(ws, width=115):
    ws.column_dimensions["A"].width = width


def main():
    epoch_metrics = pd.read_csv(RESULTS_DIR / "epoch_metrics.csv")
    seed_summary = pd.read_csv(RESULTS_DIR / "seed_summary.csv")
    per_seed_deltas = pd.read_csv(STATS_DIR / "per_seed_deltas.csv")
    seed_level_stats = pd.read_csv(STATS_DIR / "seed_level_summary_stats.csv")
    per_deg_deltas = pd.read_csv(STATS_DIR / "per_degradation_deltas.csv")
    per_deg_summary = pd.read_csv(STATS_DIR / "per_degradation_summary.csv", header=[0, 1], index_col=[0, 1])
    coeff_df = pd.read_csv(STATS_DIR / "coefficient_analysis.csv")
    control_df = pd.read_csv(STATS_DIR / "content_shuffle_controls.csv")
    variance_df = pd.read_csv(STATS_DIR / "haze_scene_variance.csv")
    rank_df = pd.read_csv(STATS_DIR / "effective_rank.csv")
    probe_df = pd.read_csv(STATS_DIR / "representation_probe.csv")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----
    ws = wb.create_sheet("README")
    autosize_text_sheet(ws)
    control_means = control_df[["psnr_normal", "psnr_degradation_only", "psnr_content_only",
                                 "psnr_shuffled_content"]].mean()
    readme_lines = [
        "TEST12: Feature-Conditioned Low-Rank Operator",
        "",
        "Research question: TEST11 showed increasing rank (2->16) of the low-rank conditional "
        "operator did NOT improve restoration -- effective rank stayed near 1-3 regardless of "
        "configured capacity. TEST12 asks a different question: does the coefficient generator "
        "need to see current spatial CONTENT (phi(F) = GAP+GMP of the bottleneck feature), not "
        "just the degradation embedding e_D? a = G([e_D; phi(F)]) vs a = G(e_D), same rank=2.",
        "",
        "HEADLINE FINDING: strong, clean causal evidence that the feature-conditioned operator "
        f"genuinely uses image content. Normal PSNR={control_means['psnr_normal']:.2f}dB; "
        f"content-only (e_D zeroed, real phi(F)) PSNR={control_means['psnr_content_only']:.2f}dB "
        "(almost identical to normal -- most of the useful signal comes from content, not the "
        f"degradation embedding alone); degradation-only (real e_D, dataset-mean phi_bar) "
        f"PSNR={control_means['psnr_degradation_only']:.2f}dB (-1.36dB from normal); shuffled "
        f"content (real e_D, WRONG image's phi(F)) PSNR={control_means['psnr_shuffled_content']:.2f}dB "
        "(worst of all conditions, -2.91dB from normal -- using mismatched content actively hurts "
        "more than using no specific content at all). T12's coefficients also vary 3.7x-11.6x more "
        "across scenes (within one degradation) than F2's, confirming F2 structurally cannot access "
        "this scene-specific signal.",
        "",
        "Restoration outcome is more modest: T12-F2 overall PSNR +0.078dB (same-sign 2/3 seeds), "
        "SSIM +0.0056 (same-sign 3/3 seeds). Haze shows the largest per-degradation movement "
        "(+0.186dB, 2/3 seeds positive with the two positive seeds much larger in magnitude than "
        "the one negative). T12-A (-0.199dB) is notably closer to baseline than F2-A (-0.276dB, "
        "consistently negative 3/3 seeds) -- content conditioning measurably narrows, but does not "
        "close, the remaining gap.",
        "",
        "Directory isolation: reuses TEST07-B's dataset/KD-cache and TEST09/11's compact-latent + "
        "low-rank mechanism definition, READ-ONLY. Imports fyp-adair-distill's locked NAFNet "
        "READ-ONLY. Does not modify any prior experiment directory.",
    ]
    for i, line in enumerate(readme_lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        if i == 1:
            c.font = Font(bold=True, size=14)

    # ---- Models ----
    ws = wb.create_sheet("Models")
    models_df = pd.DataFrame({
        "model": ["A", "F2", "T12"],
        "description": ["Baseline locked NAFNet, no KD, no conditioning",
                         "Compact latent KD + rank-2 low-rank channel mixing, a=G(e_D) (= TEST09/11's F2)",
                         "Same rank=2 operator, but a=G([e_D; phi(F)]) -- coefficient head also sees "
                         "GAP+GMP(bottleneck) content, via a 528->32->2 MLP"],
    })
    write_df(ws, models_df)

    # ---- Dataset ----
    ws = wb.create_sheet("Dataset")
    dataset_rows = pd.DataFrame({
        "field": ["source", "n_train_scenes", "n_val_scenes", "crops_per_train_scene", "crop_size_px",
                  "degradations"],
        "value": ["REUSED from test07_b/ (READ-ONLY)", 80, 20, 8, 128, "Rain, Haze, Noise"],
    })
    write_df(ws, dataset_rows)

    # ---- Training_Config ----
    ws = wb.create_sheet("Training_Config")
    config_rows = pd.DataFrame({
        "field": ["epochs", "batch_size", "learning_rate", "optimizer", "lambda_kd", "seeds", "rank",
                  "coefficient_head_T12", "coefficient_head_F2"],
        "value": [50, 8, "2e-4", "Adam", 0.1, "0, 1, 2", 2,
                  "Linear(528,32) -> ReLU -> Linear(32,2), final layer zero-init",
                  "Linear(16,2), zero-init"],
    })
    write_df(ws, config_rows)

    # ---- Epoch_Metrics ----
    ws = wb.create_sheet("Epoch_Metrics")
    write_df(ws, epoch_metrics)

    # ---- Seed_Summary ----
    ws = wb.create_sheet("Seed_Summary")
    write_df(ws, seed_summary)

    # ---- Restoration ----
    ws = wb.create_sheet("Restoration")
    write_df(ws, per_seed_deltas)
    start2 = len(per_seed_deltas) + 3
    ws.cell(row=start2, column=1, value="Seed-level summary (mean +- std, bootstrap 95% CI, same-sign count)").font = HEADER_FONT
    write_df(ws, seed_level_stats, start_row=start2 + 1)

    # ---- Per_Degradation ----
    ws = wb.create_sheet("Per_Degradation")
    write_df(ws, per_deg_deltas)
    start2 = len(per_deg_deltas) + 3
    ws.cell(row=start2, column=1, value="Per-degradation summary (mean +- std across 3 seeds)").font = HEADER_FONT
    per_deg_summary_flat = per_deg_summary.reset_index()
    per_deg_summary_flat.columns = ["comparison", "degradation", "delta_psnr_mean", "delta_psnr_std",
                                     "delta_ssim_mean", "delta_ssim_std"]
    write_df(ws, per_deg_summary_flat, start_row=start2 + 1)

    # ---- Coefficient_Analysis ----
    ws = wb.create_sheet("Coefficient_Analysis")
    write_df(ws, coeff_df)

    # ---- Content_Control ----
    ws = wb.create_sheet("Content_Control")
    write_df(ws, control_df)
    start2 = len(control_df) + 3
    ws.cell(row=start2, column=1, value="Mean PSNR by condition (across all val crops, all seeds)").font = HEADER_FONT
    cond_summary = pd.DataFrame({
        "condition": ["normal", "degradation_only", "content_only", "shuffled_content"],
        "mean_psnr": [control_df.psnr_normal.mean(), control_df.psnr_degradation_only.mean(),
                      control_df.psnr_content_only.mean(), control_df.psnr_shuffled_content.mean()],
    })
    write_df(ws, cond_summary, start_row=start2 + 1)

    # ---- Shuffle_Control ---- (subset view of the same file, per spec's separate sheet)
    ws = wb.create_sheet("Shuffle_Control")
    shuffle_cols = ["seed", "scene_id", "degradation", "psnr_normal", "psnr_shuffled_content",
                     "donor_scene_id", "donor_degradation"]
    write_df(ws, control_df[shuffle_cols])
    start2 = len(control_df) + 3
    delta = control_df.psnr_shuffled_content - control_df.psnr_normal
    ws.cell(row=start2, column=1, value="Delta PSNR (shuffled - normal): mean, std").font = HEADER_FONT
    write_df(ws, pd.DataFrame({"mean_delta": [delta.mean()], "std_delta": [delta.std()]}), start_row=start2 + 1)

    # ---- Effective_Rank ----
    ws = wb.create_sheet("Effective_Rank")
    write_df(ws, rank_df)
    start2 = len(rank_df) + 3
    ws.cell(row=start2, column=1, value="Haze scene-variance comparison (T12 vs F2)").font = HEADER_FONT
    write_df(ws, variance_df, start_row=start2 + 1)

    # ---- Complexity ----
    ws = wb.create_sheet("Complexity")
    a_row = seed_summary[seed_summary.model == "A"].iloc[0]
    f2_row = seed_summary[seed_summary.model == "F2"].iloc[0]
    t12_row = seed_summary[seed_summary.model == "T12"].iloc[0]
    complexity_rows = pd.DataFrame({
        "field": ["params_A", "params_F2", "params_T12", "extra_params_T12_vs_F2", "extra_params_T12_vs_A",
                  "macs_A", "macs_F2", "macs_T12", "note"],
        "value": [int(a_row.params), int(f2_row.params), int(t12_row.params),
                  int(t12_row.params) - int(f2_row.params), int(t12_row.params) - int(a_row.params),
                  int(a_row.macs), int(f2_row.macs), int(t12_row.macs),
                  "THEORETICAL COMPLEXITY ONLY, not an NPU latency claim."],
    })
    write_df(ws, complexity_rows)

    # ---- Representation ----
    ws = wb.create_sheet("Representation")
    write_df(ws, probe_df)

    # ---- GO_NO_GO ----
    ws = wb.create_sheet("GO_NO_GO")
    t12f2_psnr = seed_level_stats[(seed_level_stats.comparison == "T12-F2") &
                                   (seed_level_stats.metric == "delta_last5_psnr")].iloc[0]
    haze_t12f2 = per_deg_summary.loc[("T12-F2", "haze"), ("delta_psnr", "mean")]
    shuffle_delta = (control_df.psnr_shuffled_content - control_df.psnr_normal).mean()
    content_delta = (control_df.psnr_content_only - control_df.psnr_normal).mean()
    go_rows = pd.DataFrame({
        "field": ["decision", "T12_minus_F2_mean_psnr", "T12_minus_F2_same_sign_psnr",
                  "haze_T12_minus_F2_mean_psnr", "content_only_vs_normal_delta",
                  "shuffled_vs_normal_delta", "scene_variance_ratio_haze", "operator_uses_content",
                  "rationale"],
        "value": ["PARTIAL GO", round(t12f2_psnr["mean"], 4), f"{t12f2_psnr['same_sign_count']}/3",
                  round(haze_t12f2, 4), round(content_delta, 3), round(shuffle_delta, 3),
                  round(variance_df[variance_df.degradation == "Haze"].T12_over_F2_variance_ratio_a0.mean(), 2),
                  True,
                  "Strong causal evidence (GO-level) that the operator genuinely uses spatial "
                  "content: content-only control (e_D zeroed) performs almost identically to "
                  "normal, while shuffled content (wrong image) is the WORST of all four "
                  "conditions -- worse than using no specific content at all. T12's coefficients "
                  "vary 3.7-11.6x more across scenes than F2's within the same degradation, "
                  "confirming F2 structurally cannot access this signal. However, the resulting "
                  "restoration gain is modest and only partially consistent: overall T12-F2 PSNR "
                  "is +0.078dB but only 2/3 seeds share the same sign (SSIM is 3/3 consistent, "
                  "smaller magnitude). Haze shows the largest per-degradation movement (+0.186dB, "
                  "2/3 seeds positive with much larger magnitude on the positive seeds) but is not "
                  "unanimous either. This matches PARTIAL GO: overall gain small/inconsistent, but "
                  "Haze shows the most meaningful movement of the three degradations, combined "
                  "with unambiguous causal evidence the mechanism itself works as hypothesized."],
    })
    write_df(ws, go_rows)

    # ---- Visualizations ----
    ws = wb.create_sheet("Visualizations")
    row_cursor = 1
    for png in sorted(VIZ_DIR.glob("*.png")):
        ws.cell(row=row_cursor, column=1, value=png.stem).font = HEADER_FONT
        img = XLImage(str(png))
        img.width, img.height = 480, 300
        ws.add_image(img, f"A{row_cursor + 1}")
        row_cursor += 18

    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")


if __name__ == "__main__":
    main()
